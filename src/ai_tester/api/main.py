"""
FastAPI backend for AI Tester Framework.
Provides REST API and WebSocket endpoints for the web UI.
"""
from fastapi import FastAPI, HTTPException, WebSocket, WebSocketDisconnect, UploadFile, File, Form, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, JSONResponse
from fastapi.exceptions import RequestValidationError
from pydantic import BaseModel, ValidationError
from typing import List, Optional, Dict, Any
import asyncio
import json
import io
import csv
import os
from datetime import datetime
from dotenv import load_dotenv
from cachetools import TTLCache

# Load environment variables from .env file
load_dotenv()

from ai_tester.clients.jira_client import JiraClient
from ai_tester.clients.llm_client import LLMClient
from ai_tester.agents.strategic_planner import StrategicPlannerAgent
from ai_tester.agents.evaluator import EvaluationAgent
from ai_tester.agents.ticket_analyzer import TicketAnalyzerAgent
from ai_tester.agents.test_ticket_generator import TestTicketGeneratorAgent
from ai_tester.agents.test_ticket_reviewer import TestTicketReviewerAgent
from ai_tester.agents.questioner_agent import QuestionerAgent
from ai_tester.agents.gap_analyzer_agent import GapAnalyzerAgent
from ai_tester.agents.ticket_improver_agent import TicketImproverAgent
from ai_tester.agents.coverage_reviewer_agent import CoverageReviewerAgent
from ai_tester.agents.requirements_fixer_agent import RequirementsFixerAgent
from ai_tester.core.models import TestCase, TestStep
from ai_tester.core.test_ticket_models import TestTicket
from ai_tester.utils.jira_text_cleaner import clean_jira_text_for_llm
from ai_tester.utils.utils import adf_to_plaintext
import re

# Validation utilities
def validate_jira_key(key: str) -> str:
    """
    Validate Jira key format (PROJECT-123).

    Args:
        key: The Jira key to validate

    Returns:
        The validated key

    Raises:
        HTTPException: If key format is invalid
    """
    if not key:
        raise HTTPException(status_code=400, detail="Jira key cannot be empty")

    # Jira key format: PROJECT-NUMBER (e.g., PFI-1848)
    # Project key: 1-10 uppercase letters/numbers, must start with letter
    # Issue number: 1-10 digits
    if not re.match(r'^[A-Z][A-Z0-9]{0,9}-\d{1,10}$', key):
        raise HTTPException(
            status_code=400,
            detail=f"Invalid Jira key format: '{key}'. Expected format: PROJECT-123"
        )

    if len(key) > 50:
        raise HTTPException(status_code=400, detail="Jira key too long (max 50 characters)")

    return key

# File upload limits and validation
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10 MB per file
MAX_FILES = 5  # Maximum 5 files per request
ALLOWED_MIME_TYPES = {
    'application/pdf',
    'application/vnd.openxmlformats-officedocument.wordprocessingml.document',  # .docx
    'application/msword',  # .doc
    'text/plain',
    'text/markdown',
    'image/png',
    'image/jpeg',
    'image/jpg',
    'image/gif'
}
ALLOWED_EXTENSIONS = {'.pdf', '.docx', '.doc', '.txt', '.md', '.png', '.jpg', '.jpeg', '.gif'}

async def validate_uploaded_files(files: List[UploadFile]) -> None:
    """
    Validate uploaded files for size, count, and type.

    Args:
        files: List of uploaded files

    Raises:
        HTTPException: If validation fails
    """
    if not files:
        return

    # Check file count
    if len(files) > MAX_FILES:
        raise HTTPException(
            status_code=400,
            detail=f"Too many files. Maximum {MAX_FILES} files allowed, got {len(files)}"
        )

    total_size = 0

    for file in files:
        # Check individual file size by reading in chunks
        file_size = 0
        chunk_size = 1024 * 1024  # 1MB chunks

        while chunk := await file.read(chunk_size):
            file_size += len(chunk)
            if file_size > MAX_FILE_SIZE:
                raise HTTPException(
                    status_code=413,
                    detail=f"File '{file.filename}' exceeds maximum size of {MAX_FILE_SIZE // (1024*1024)}MB"
                )

        # Reset file pointer after reading
        await file.seek(0)

        total_size += file_size

        # Check total size across all files (50MB total)
        if total_size > MAX_FILE_SIZE * 5:
            raise HTTPException(
                status_code=413,
                detail=f"Total upload size exceeds maximum of {(MAX_FILE_SIZE * 5) // (1024*1024)}MB"
            )

        # Validate file extension
        if file.filename:
            ext = '.' + file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
            if ext not in ALLOWED_EXTENSIONS:
                raise HTTPException(
                    status_code=400,
                    detail=f"File type not allowed: '{ext}'. Allowed types: {', '.join(ALLOWED_EXTENSIONS)}"
                )

        # Validate MIME type
        if file.content_type and file.content_type not in ALLOWED_MIME_TYPES:
            raise HTTPException(
                status_code=400,
                detail=f"MIME type not allowed: '{file.content_type}'. Allowed types: PDF, Word, text, images"
            )

# Initialize FastAPI app
app = FastAPI(
    title="AI Tester Framework API",
    description="Backend API for AI-powered test case and test ticket generation",
    version="3.0.0"
)

# CORS middleware for React frontend
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000", "http://localhost:5173"],  # React dev servers
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Validation error handler for better debugging
@app.exception_handler(RequestValidationError)
async def validation_exception_handler(request: Request, exc: RequestValidationError):
    body = await request.body()
    print(f"DEBUG Validation Error: {exc}")
    print(f"DEBUG Request body: {body}")
    print(f"DEBUG Validation errors: {exc.errors()}")
    return JSONResponse(
        status_code=422,
        content={"detail": exc.errors(), "body": body.decode() if body else "empty"}
    )

# Global clients (will be initialized with credentials)
jira_client: Optional[JiraClient] = None
llm_client: Optional[LLMClient] = None
client_init_lock = asyncio.Lock()  # Protects client initialization from race conditions

# In-memory storage for generated test tickets (no TTL - user manages lifecycle)
test_tickets_storage: Dict[str, TestTicket] = {}
test_tickets_lock = asyncio.Lock()

# In-memory cache for improved tickets (ticket_key -> improvement_data)
# TTL: 1 hour, Max size: 1000 tickets
# This avoids duplicate LLM calls for the same ticket
improved_tickets_cache = TTLCache(maxsize=1000, ttl=3600)
improved_tickets_lock = asyncio.Lock()

# In-memory cache for epic attachments (epic_key -> {epic_attachments, child_attachments})
# TTL: 2 hours, Max size: 100 epics
# This preserves uploaded documents across Epic Analysis -> Test Ticket Generation requests
epic_attachments_cache = TTLCache(maxsize=100, ttl=7200)
epic_attachments_lock = asyncio.Lock()

# In-memory storage for user settings
user_settings: Dict[str, Any] = {
    "multiAgentMode": True,
    "maxIterations": 3,
    "qualityThreshold": 80,
    "autoValidation": True,
    "enableCriticAgent": True,
    "enableRefinement": True
}

# WebSocket connection manager
class ConnectionManager:
    def __init__(self):
        self.active_connections: List[WebSocket] = []

    async def connect(self, websocket: WebSocket):
        await websocket.accept()
        self.active_connections.append(websocket)

    def disconnect(self, websocket: WebSocket):
        self.active_connections.remove(websocket)

    async def send_progress(self, message: dict):
        """Send progress update to all connected clients."""
        for connection in self.active_connections:
            try:
                await connection.send_json(message)
            except:
                pass

manager = ConnectionManager()


# ============================================================================
# Request/Response Models
# ============================================================================

class JiraCredentials(BaseModel):
    base_url: str
    email: str
    api_token: str

class TicketKey(BaseModel):
    key: str

class EpicAnalysisRequest(BaseModel):
    epic_key: str
    include_attachments: bool = True

class TestCaseGenerationRequest(BaseModel):
    ticket_key: str
    include_attachments: bool = True

class TestTicketGenerationRequest(BaseModel):
    epic_key: str
    selected_option_index: Optional[int] = None
    selected_option: Optional[Dict[str, Any]] = None  # Pass the actual option data to avoid re-analysis

class StrategicOption(BaseModel):
    strategy: str
    description: str
    test_tickets: List[Dict[str, Any]]
    rationale: str
    advantages: List[str]
    disadvantages: List[str]
    evaluation: Optional[Dict[str, Any]] = None

class TestCaseResponse(BaseModel):
    test_cases: List[Dict[str, Any]]
    ticket_info: Dict[str, Any]
    requirements: List[Dict[str, Any]] = []
    generated_at: str
    improved_ticket: Optional[Dict[str, Any]] = None  # Preprocessed ticket for analysis

class ExportRequest(BaseModel):
    test_cases: List[Dict[str, Any]]
    ticket_key: str
    format: str  # "csv", "xlsx", or "testrail"


# ============================================================================
# Authentication & Setup
# ============================================================================

@app.post("/api/auth/login")
async def login(credentials: JiraCredentials):
    """Initialize Jira and LLM clients with credentials."""
    global jira_client, llm_client

    print("\n" + "="*80)
    print("DEBUG: Login attempt starting")
    print(f"DEBUG: Base URL: {credentials.base_url}")
    print(f"DEBUG: Email: {credentials.email}")
    print(f"DEBUG: API Token length: {len(credentials.api_token)} characters")
    print(f"DEBUG: API Token starts with: {credentials.api_token[:4]}..." if credentials.api_token else "DEBUG: No API token provided")
    print("="*80 + "\n")

    try:
        # Use lock to prevent race conditions during concurrent login attempts
        async with client_init_lock:
            print("DEBUG: Initializing JiraClient...")
            # Initialize clients
            jira_client = JiraClient(
                base_url=credentials.base_url,
                email=credentials.email,
                api_token=credentials.api_token,
                enable_pii_detection=True  # Phase 2.2: Enable PII detection and entity pseudonymization
            )
            print("DEBUG: JiraClient initialized successfully")

            print("DEBUG: Initializing LLMClient...")
            llm_client = LLMClient()
            print("DEBUG: LLMClient initialized successfully")

            # Test connection by making a simple search request
            # This validates credentials without requiring a specific issue
            print("DEBUG: Testing Jira connection with search_jql...")
            try:
                jira_client.search_jql("", ["key"], max_results=1)
                print("DEBUG: Jira connection test successful!")
            except Exception as jql_error:
                print(f"DEBUG: Jira connection test FAILED: {str(jql_error)}")
                print(f"DEBUG: Error type: {type(jql_error).__name__}")
                raise

        print("DEBUG: Authentication successful - returning success response\n")
        return {
            "success": True,
            "message": "Successfully authenticated with Jira",
            "jira_url": credentials.base_url
        }
    except Exception as e:
        print("\n" + "="*80)
        print("DEBUG: Authentication FAILED")
        print(f"DEBUG: Exception type: {type(e).__name__}")
        print(f"DEBUG: Exception message: {str(e)}")
        print(f"DEBUG: Exception args: {e.args}")

        # Print full traceback for debugging
        import traceback
        print("DEBUG: Full traceback:")
        print(traceback.format_exc())
        print("="*80 + "\n")

        error_message = str(e).lower()

        # Parse specific error types
        if "401" in error_message or "unauthorized" in error_message:
            detail = "Invalid credentials. Please check your email and API token."
        elif "403" in error_message or "forbidden" in error_message:
            detail = "Access denied. Your account may not have permission to access this Jira instance."
        elif "404" in error_message or "not found" in error_message:
            detail = "Jira instance not found. Please verify the base URL is correct."
        elif "timeout" in error_message or "timed out" in error_message:
            detail = "Connection timeout. Please check your network connection and Jira URL."
        elif "connection" in error_message or "network" in error_message:
            detail = "Cannot connect to Jira. Please verify the base URL and your internet connection."
        elif "ssl" in error_message or "certificate" in error_message:
            detail = "SSL certificate error. Please check the Jira URL uses HTTPS."
        else:
            detail = f"Authentication failed: {str(e)}"

        print(f"DEBUG: Sending error response: {detail}\n")
        raise HTTPException(status_code=401, detail=detail)


@app.get("/api/auth/status")
async def auth_status():
    """Check if clients are initialized."""
    is_jira_initialized = jira_client is not None
    is_llm_initialized = llm_client is not None
    is_authenticated = is_jira_initialized and is_llm_initialized

    print("\n" + "-"*80)
    print("DEBUG: Auth status check")
    print(f"DEBUG: jira_client exists: {is_jira_initialized}")
    print(f"DEBUG: llm_client exists: {is_llm_initialized}")
    print(f"DEBUG: Is authenticated: {is_authenticated}")
    if jira_client:
        print(f"DEBUG: Jira base URL: {jira_client.base_url}")
    print("-"*80 + "\n")

    return {
        "authenticated": is_authenticated,
        "jira_url": jira_client.base_url if jira_client else None
    }


# ============================================================================
# Epic & Ticket Loading
# ============================================================================

@app.post("/api/epics/load")
async def load_epic(request: EpicAnalysisRequest):
    """Load Epic from Jira with all children and attachments."""
    if not jira_client:
        raise HTTPException(status_code=401, detail="Not authenticated")

    await manager.send_progress({
        "type": "progress",
        "step": "loading_epic",
        "message": f"Loading Epic {request.epic_key}..."
    })

    try:
        # Fetch Epic
        epic = jira_client.get_issue(request.epic_key)

        await manager.send_progress({
            "type": "progress",
            "step": "loading_children",
            "message": "Loading child tickets..."
        })

        # Fetch children
        children = jira_client.get_children_of_epic(request.epic_key)

        # Process attachments if requested
        epic_attachments = []
        child_attachments = {}

        if request.include_attachments:
            await manager.send_progress({
                "type": "progress",
                "step": "loading_attachments",
                "message": "Processing epic attachments..."
            })

            # Process epic attachments
            epic_attachment_list = jira_client.get_attachments(request.epic_key)
            for attachment in epic_attachment_list:
                processed = jira_client.process_attachment(attachment)
                if processed:
                    epic_attachments.append(processed)

            print(f"DEBUG: Processed {len(epic_attachments)} attachments for epic {request.epic_key}")

            # Also extract attachments referenced in the description (embedded media)
            print(f"DEBUG: Checking for attachments embedded in description (epic load)")
            epic_desc_field = epic.get('fields', {}).get('description')
            if epic_desc_field and isinstance(epic_desc_field, dict):
                embedded_attachments = jira_client.extract_attachments_from_description(request.epic_key, epic_desc_field)
                print(f"DEBUG: Found {len(embedded_attachments)} attachments embedded in description")

                # Process any embedded attachments that aren't already in the list
                already_processed = {att.get('filename') for att in epic_attachments}
                for attachment in embedded_attachments:
                    filename = attachment.get('filename')
                    if filename not in already_processed:
                        print(f"DEBUG: Processing embedded attachment: {filename}")
                        processed = jira_client.process_attachment(attachment)
                        if processed:
                            epic_attachments.append(processed)
                            print(f"DEBUG: Successfully processed embedded: {processed.get('filename')}")
                    else:
                        print(f"DEBUG: Embedded attachment {filename} already processed")

            print(f"DEBUG: Total epic attachments after checking description: {len(epic_attachments)}")

            # Process child ticket attachments
            for idx, child in enumerate(children):
                child_key = child.get('key')
                if child_key:
                    await manager.send_progress({
                        "type": "progress",
                        "step": "loading_child_attachments",
                        "message": f"Processing attachments for {child_key} ({idx+1}/{len(children)})..."
                    })

                    child_attachment_list = jira_client.get_attachments(child_key)
                    processed_child_attachments = []
                    for attachment in child_attachment_list:
                        processed = jira_client.process_attachment(attachment)
                        if processed:
                            processed_child_attachments.append(processed)

                    if processed_child_attachments:
                        child_attachments[child_key] = processed_child_attachments
                        print(f"DEBUG: Processed {len(processed_child_attachments)} attachments for {child_key}")

        await manager.send_progress({
            "type": "complete",
            "message": "Epic loaded successfully"
        })

        return {
            "epic": epic,
            "children": children,
            "child_count": len(children),
            "epic_attachments": epic_attachments,
            "child_attachments": child_attachments
        }

    except Exception as e:
        await manager.send_progress({
            "type": "error",
            "message": str(e)
        })
        raise HTTPException(status_code=500, detail=str(e))


class ManualTicketsRequest(BaseModel):
    ticket_keys: List[str]


@app.post("/api/epics/{epic_key}/add-children")
async def add_children_to_epic(epic_key: str, request: ManualTicketsRequest):
    """
    Manually add child tickets to an Epic/Initiative.

    This is a fallback method for when automatic child loading fails,
    particularly useful for Initiatives where Jira's API may not return
    all child Epics properly.

    Args:
        epic_key: The Epic/Initiative key
        request: List of ticket keys to add as children

    Returns:
        List of loaded tickets with metadata
    """
    if not jira_client:
        raise HTTPException(status_code=401, detail="Not authenticated")

    # Validate epic key and all child ticket keys
    epic_key = validate_jira_key(epic_key)
    for ticket_key in request.ticket_keys:
        validate_jira_key(ticket_key)

    await manager.send_progress({
        "type": "progress",
        "step": "loading_manual_children",
        "message": f"Loading {len(request.ticket_keys)} tickets..."
    })

    try:
        loaded_tickets = []
        failed_tickets = []

        for ticket_key in request.ticket_keys:
            try:
                await manager.send_progress({
                    "type": "progress",
                    "step": "loading_ticket",
                    "message": f"Loading {ticket_key}..."
                })

                ticket = jira_client.get_issue(ticket_key)
                loaded_tickets.append(ticket)

            except Exception as e:
                print(f"Failed to load {ticket_key}: {str(e)}")
                failed_tickets.append({
                    "key": ticket_key,
                    "error": str(e)
                })

        await manager.send_progress({
            "type": "complete",
            "message": f"Loaded {len(loaded_tickets)} tickets successfully"
        })

        return {
            "success": True,
            "loaded_tickets": loaded_tickets,
            "loaded_count": len(loaded_tickets),
            "failed_tickets": failed_tickets,
            "failed_count": len(failed_tickets)
        }

    except Exception as e:
        await manager.send_progress({
            "type": "error",
            "message": str(e)
        })
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/tickets/{ticket_key}")
async def get_ticket(ticket_key: str):
    """Get a single Jira ticket by key."""
    if not jira_client:
        raise HTTPException(status_code=401, detail="Not authenticated")

    # Validate Jira key format
    ticket_key = validate_jira_key(ticket_key)

    try:
        print(f"DEBUG: Fetching ticket {ticket_key} from Jira...")
        ticket = jira_client.get_issue(ticket_key)
        print(f"DEBUG: Ticket fetched successfully")

        # Extract fields
        fields = ticket.get("fields", {})

        # Convert description from ADF to plaintext if needed
        description = fields.get('description', '')
        if isinstance(description, dict):
            from ai_tester.utils.utils import adf_to_plaintext
            description_plaintext = adf_to_plaintext(description)
            fields['description'] = description_plaintext
            ticket['description_raw'] = description
        elif description is None:
            fields['description'] = ''

        # Extract acceptance criteria from custom fields
        acceptance_criteria = ""
        acceptance_criteria_raw = None

        # First try the known custom field ID for this Jira instance
        ac_field_value = fields.get("customfield_10524")
        if ac_field_value:
            if isinstance(ac_field_value, str):
                acceptance_criteria = ac_field_value
                acceptance_criteria_raw = ac_field_value
            elif isinstance(ac_field_value, dict):
                from ai_tester.utils.utils import adf_to_plaintext
                acceptance_criteria = adf_to_plaintext(ac_field_value)
                acceptance_criteria_raw = ac_field_value

        # Fallback: Look for acceptance criteria in custom fields by name
        if not acceptance_criteria:
            for field_key, field_value in fields.items():
                if "acceptance" in field_key.lower() or "criteria" in field_key.lower():
                    if field_value:
                        if isinstance(field_value, str):
                            acceptance_criteria = field_value
                            acceptance_criteria_raw = field_value
                        elif isinstance(field_value, dict):
                            from ai_tester.utils.utils import adf_to_plaintext
                            acceptance_criteria = adf_to_plaintext(field_value)
                            acceptance_criteria_raw = field_value
                        break

        # Add extracted acceptance criteria to response
        if acceptance_criteria:
            ticket['acceptance_criteria'] = acceptance_criteria
            ticket['acceptance_criteria_raw'] = acceptance_criteria_raw
            print(f"Added acceptance_criteria to ticket response")
        else:
            print(f"No acceptance criteria found for ticket {ticket.get('key')}")

        return ticket
    except Exception as e:
        import traceback
        import sys
        error_trace = traceback.format_exc()
        # Write to stderr which uvicorn should capture
        print(f"\n===ERROR fetching ticket {ticket_key}===", file=sys.stderr)
        print(error_trace, file=sys.stderr)
        raise HTTPException(status_code=404, detail=f"Ticket not found: {str(e)}")


@app.post("/api/tickets/{ticket_key}/analyze")
async def analyze_ticket(ticket_key: str):
    """Analyze a ticket for test case generation readiness."""
    if not jira_client or not llm_client:
        raise HTTPException(status_code=401, detail="Not authenticated")

    # Validate Jira key format
    ticket_key = validate_jira_key(ticket_key)

    await manager.send_progress({
        "type": "progress",
        "step": "loading_ticket",
        "message": f"Loading ticket {ticket_key}..."
    })

    try:
        # Load ticket
        ticket = jira_client.get_issue(ticket_key)
        fields = ticket.get('fields', {})

        # Process ticket description (convert from ADF if needed)
        ticket_description = fields.get('description', '')
        if isinstance(ticket_description, dict):
            from ai_tester.utils.utils import adf_to_plaintext
            ticket_description = adf_to_plaintext(ticket_description)
        elif ticket_description is None:
            ticket_description = ''

        # Extract acceptance criteria from custom fields
        acceptance_criteria = ""
        for field_key, field_value in fields.items():
            if "acceptance" in field_key.lower() or "criteria" in field_key.lower():
                if field_value:
                    if isinstance(field_value, str):
                        acceptance_criteria = field_value
                    elif isinstance(field_value, dict):
                        from ai_tester.utils.utils import adf_to_plaintext
                        acceptance_criteria = adf_to_plaintext(field_value)
                    break

        # Create ticket data structure for agents
        # Include both description and acceptance criteria
        full_description = ticket_description
        if acceptance_criteria:
            full_description += f"\n\n=== ACCEPTANCE CRITERIA ===\n{acceptance_criteria}"

        ticket_data = {
            'key': ticket.get('key', ''),
            'summary': fields.get('summary', ''),
            'description': full_description,
            'acceptance_criteria': acceptance_criteria
        }

        await manager.send_progress({
            "type": "progress",
            "step": "analyzing",
            "message": "Generating questions about ticket gaps..."
        })

        # Use QuestionerAgent to generate questions
        from ai_tester.agents.questioner_agent import QuestionerAgent
        questioner = QuestionerAgent(llm_client)

        # QuestionerAgent expects epic_data and child_tickets, but for single ticket analysis
        # we pass the ticket as epic_data and empty child_tickets
        questions, error = await asyncio.to_thread(
            questioner.generate_questions,
            ticket_data,
            []  # No child tickets for single ticket analysis
        )

        if error:
            raise Exception(f"Failed to generate questions: {error}")

        await manager.send_progress({
            "type": "progress",
            "step": "prioritizing",
            "message": "Prioritizing questions and assessing readiness..."
        })

        # Use GapAnalyzerAgent to prioritize questions
        from ai_tester.agents.gap_analyzer_agent import GapAnalyzerAgent
        gap_analyzer = GapAnalyzerAgent(llm_client)

        assessment, error = await asyncio.to_thread(
            gap_analyzer.analyze_questions,
            questions,
            ticket_data,
            []  # No child tickets for single ticket analysis
        )

        if error:
            raise Exception(f"Failed to analyze gaps: {error}")

        await manager.send_progress({
            "type": "complete",
            "message": "Analysis complete"
        })

        return {
            "ticket_key": ticket_key,
            "assessment": assessment,
            "analyzed_at": datetime.now().isoformat()
        }

    except Exception as e:
        await manager.send_progress({
            "type": "error",
            "message": str(e)
        })
        raise HTTPException(status_code=500, detail=str(e))


# ============================================================================
# Multi-Agent Test Ticket Generation
# ============================================================================

@app.post("/api/epics/{epic_key}/analyze")
async def analyze_epic(
    epic_key: str,
    files: List[UploadFile] = File(default=[])
):
    """
    Use Strategic Planner to generate 3 strategic options for splitting an Epic.
    Then use Evaluator to score each option.
    Optionally accepts uploaded supporting documents (PDFs, Word, images, markdown).
    """
    if not jira_client or not llm_client:
        raise HTTPException(status_code=401, detail="Not authenticated")

    # Validate Jira key format
    epic_key = validate_jira_key(epic_key)

    # Validate uploaded files
    await validate_uploaded_files(files)

    await manager.send_progress({
        "type": "progress",
        "step": "loading_epic",
        "message": f"Loading Epic {epic_key}..."
    })

    try:
        # Clear any existing cached attachments for this epic to ensure fresh analysis
        # This prevents stale uploaded documents from a previous analysis being reused
        async with epic_attachments_lock:
            if epic_key in epic_attachments_cache:
                print(f"DEBUG: Clearing existing cached attachments for {epic_key}")
                del epic_attachments_cache[epic_key]

        # Load Epic and children
        epic = jira_client.get_issue(epic_key)
        children_raw = jira_client.get_children_of_epic(epic_key)

        await manager.send_progress({
            "type": "progress",
            "step": "strategic_planning",
            "message": "Strategic Planner is analyzing the Epic..."
        })

        # Initialize agents
        planner = StrategicPlannerAgent(llm_client)
        evaluator = EvaluationAgent(llm_client)

        # Generate strategic options
        # Build context for strategic planner
        
        # Process epic description (convert from ADF if needed)
        epic_description = epic.get('fields', {}).get('description', '')
        if isinstance(epic_description, dict):
            from ai_tester.utils.utils import adf_to_plaintext
            epic_description = adf_to_plaintext(epic_description)
        elif epic_description is None:
            epic_description = ''
        
        # Transform children to expected format and convert descriptions
        children = []
        for child in children_raw:
            fields = child.get('fields', {})
            desc = fields.get('description', '')

            # Convert ADF description to plaintext if needed
            if isinstance(desc, dict):
                from ai_tester.utils.utils import adf_to_plaintext
                desc = adf_to_plaintext(desc)
            elif desc is None:
                desc = ''

            children.append({
                'key': child.get('key', ''),
                'summary': fields.get('summary', ''),
                'desc': desc
            })

        # Load and process attachments
        print(f"DEBUG: Starting attachment processing for epic {epic_key}")
        await manager.send_progress({
            "type": "progress",
            "step": "loading_attachments",
            "message": "Processing epic attachments..."
        })

        epic_attachments = []
        child_attachments = {}

        # Process epic attachments
        print(f"DEBUG: Calling get_attachments for epic {epic_key}")
        epic_attachment_list = jira_client.get_attachments(epic_key)
        print(f"DEBUG: Retrieved {len(epic_attachment_list)} raw attachments for epic {epic_key}")

        for attachment in epic_attachment_list:
            print(f"DEBUG: Processing attachment: {attachment.get('filename', 'unknown')}")
            processed = jira_client.process_attachment(attachment)
            if processed:
                epic_attachments.append(processed)
                print(f"DEBUG: Successfully processed: {processed.get('filename')}")
            else:
                print(f"DEBUG: Failed to process attachment: {attachment.get('filename', 'unknown')}")

        print(f"DEBUG: Processed {len(epic_attachments)} attachments for epic {epic_key}")

        # Also extract attachments referenced in the description (embedded media)
        print(f"DEBUG: Checking for attachments embedded in description")
        epic_desc_field = epic.get('fields', {}).get('description')
        if epic_desc_field and isinstance(epic_desc_field, dict):
            embedded_attachments = jira_client.extract_attachments_from_description(epic_key, epic_desc_field)
            print(f"DEBUG: Found {len(embedded_attachments)} attachments embedded in description")

            # Process any embedded attachments that aren't already in the list
            already_processed = {att.get('filename') for att in epic_attachments}
            for attachment in embedded_attachments:
                filename = attachment.get('filename')
                if filename not in already_processed:
                    print(f"DEBUG: Processing embedded attachment: {filename}")
                    processed = jira_client.process_attachment(attachment)
                    if processed:
                        epic_attachments.append(processed)
                        print(f"DEBUG: Successfully processed embedded: {processed.get('filename')}")
                else:
                    print(f"DEBUG: Embedded attachment {filename} already processed")

        print(f"DEBUG: Total epic attachments after checking description: {len(epic_attachments)}")

        # Process uploaded documents
        if 'files' in locals() and files and len(files) > 0:
            if 'manager' in locals():
                await manager.send_progress({
                    "type": "progress",
                    "step": "processing_uploads",
                    "message": f"Processing {len(files)} uploaded documents..."
                })

            print(f"DEBUG: Processing {len(files)} uploaded documents")

            for uploaded_file in files:
                try:
                    file_bytes = await uploaded_file.read()
                    print(f"DEBUG: Processing uploaded file: {uploaded_file.filename} ({uploaded_file.content_type}, {len(file_bytes)} bytes)")

                    result = {
                        "filename": uploaded_file.filename,
                        "mime_type": uploaded_file.content_type,
                        "size": len(file_bytes)
                    }

                    # PDF files
                    if uploaded_file.content_type == "application/pdf" or uploaded_file.filename.lower().endswith('.pdf'):
                        from ai_tester.utils.utils import extract_text_from_pdf
                        text = extract_text_from_pdf(file_bytes)
                        result["type"] = "document"
                        result["content"] = text
                        epic_attachments.append(result)
                        print(f"DEBUG: Successfully processed uploaded PDF: {uploaded_file.filename}")

                    # Word documents
                    elif uploaded_file.content_type in ["application/vnd.openxmlformats-officedocument.wordprocessingml.document", "application/msword"] or uploaded_file.filename.lower().endswith(('.docx', '.doc')):
                        from ai_tester.utils.utils import extract_text_from_word
                        text = extract_text_from_word(file_bytes)
                        result["type"] = "document"
                        result["content"] = text
                        epic_attachments.append(result)
                        print(f"DEBUG: Successfully processed uploaded Word doc: {uploaded_file.filename}")

                    # Images
                    elif uploaded_file.content_type and uploaded_file.content_type.startswith("image/"):
                        if uploaded_file.content_type in ["image/jpeg", "image/png", "image/gif", "image/webp"]:
                            from ai_tester.utils.utils import encode_image_to_base64
                            base64_data = encode_image_to_base64(file_bytes, uploaded_file.content_type)
                            result["type"] = "image"
                            result["content"] = base64_data
                            result["data_url"] = f"data:{uploaded_file.content_type};base64,{base64_data}"
                            epic_attachments.append(result)
                            print(f"DEBUG: Successfully processed uploaded image: {uploaded_file.filename}")

                    # Text/Markdown files
                    elif uploaded_file.content_type and (uploaded_file.content_type.startswith("text/") or uploaded_file.filename.lower().endswith(('.txt', '.md'))):
                        try:
                            text = file_bytes.decode('utf-8')
                            result["type"] = "document"
                            result["content"] = text
                            epic_attachments.append(result)
                            print(f"DEBUG: Successfully processed uploaded text file: {uploaded_file.filename}")
                        except Exception as e:
                            print(f"DEBUG: Error decoding text file {uploaded_file.filename}: {e}")

                except Exception as e:
                    print(f"DEBUG: Error processing uploaded file {uploaded_file.filename}: {e}")

            print(f"DEBUG: Total epic attachments after processing uploads: {len(epic_attachments)}")

        # Process child ticket attachments (limit to avoid excessive API calls)
        for idx, child in enumerate(children_raw[:20]):  # Only first 20 children to avoid slowdown
            child_key = child.get('key')
            if child_key:
                child_attachment_list = jira_client.get_attachments(child_key)
                processed_child_attachments = []
                for attachment in child_attachment_list:
                    processed = jira_client.process_attachment(attachment)
                    if processed:
                        processed_child_attachments.append(processed)

                if processed_child_attachments:
                    child_attachments[child_key] = processed_child_attachments

        epic_context = {
            'epic_key': epic.get('key'),
            'epic_summary': epic.get('fields', {}).get('summary'),
            'epic_desc': epic_description,
            'children': children,
            'epic_attachments': epic_attachments,
            'child_attachments': child_attachments
        }

        # Preprocess Epic and analyze attachments IN PARALLEL for performance
        await manager.send_progress({
            "type": "progress",
            "step": "preprocessing",
            "message": "Preprocessing epic and analyzing attachments in parallel..."
        })

        improved_tickets = {}
        pre_analyzed_attachments = None

        # Define async tasks for parallel execution
        async def preprocess_epic():
            """Preprocess Epic ticket for consistency"""
            try:
                epic_key = epic.get('key')

                # Check cache first
                async with improved_tickets_lock:
                    if epic_key in improved_tickets_cache:
                        print(f"DEBUG: Using cached improved epic for {epic_key}")
                        cached_result = improved_tickets_cache[epic_key]
                        return cached_result.get("improved_ticket", {})

                from ai_tester.agents.ticket_improver_agent import TicketImproverAgent
                ticket_improver = TicketImproverAgent(llm_client)

                epic_improvement, epic_improve_error = await asyncio.to_thread(
                    ticket_improver.improve_ticket,
                    {
                        "key": epic_key,
                        "summary": epic.get('fields', {}).get('summary', ''),
                        "description": epic_description
                    },
                    None,  # questions
                    None,  # epic_context
                    'gpt-4o-mini'  # Use cheaper model for data extraction
                )

                if epic_improvement and not epic_improve_error:
                    # Cache the result
                    async with improved_tickets_lock:
                        improved_tickets_cache[epic_key] = epic_improvement
                    improved_epic = epic_improvement.get("improved_ticket", {})
                    print(f"DEBUG: Epic {epic_key} preprocessed and cached successfully")
                    return improved_epic
                return None
            except Exception as e:
                print(f"DEBUG: Epic preprocessing error: {e}")
                return None

        async def analyze_attachments():
            """Analyze attachments (images/documents) using vision API"""
            try:
                print(f"DEBUG: Starting attachment analysis for {len(epic_attachments)} attachments")
                result = await asyncio.to_thread(
                    planner.analyze_attachments,
                    epic_attachments,
                    child_attachments
                )
                print(f"DEBUG: Attachment analysis complete")
                return result
            except Exception as e:
                print(f"DEBUG: Attachment analysis error: {e}")
                return None

        # Execute both tasks in parallel using asyncio.gather
        print(f"DEBUG: Starting parallel preprocessing (Epic preprocessing + Attachment analysis)")
        improved_epic, attachment_analysis = await asyncio.gather(
            preprocess_epic(),
            analyze_attachments()
        )
        print(f"DEBUG: Parallel preprocessing complete")

        # Process results
        if improved_epic:
            improved_tickets[epic.get('key')] = improved_epic
            # Update epic context with improved description
            if improved_epic.get("description"):
                epic_context['epic_desc'] = improved_epic["description"]
            print(f"DEBUG: Applied improved epic description")

        if attachment_analysis:
            pre_analyzed_attachments = attachment_analysis
            print(f"DEBUG: Using pre-analyzed attachments: {len(attachment_analysis.get('image_analysis', {}))} images, {len(attachment_analysis.get('document_summaries', {}))} documents")

        print(f"DEBUG: Total tickets preprocessed: {len(improved_tickets)}")

        # Store improved tickets in context for frontend display
        epic_context['improved_tickets'] = improved_tickets

        # Call the run method with pre-analyzed attachments
        await manager.send_progress({
            "type": "progress",
            "step": "planning",
            "message": "Generating strategic options..."
        })

        options, error = await asyncio.to_thread(
            planner.propose_splits,
            epic_context,
            pre_analyzed_attachments
        )

        if error:
            raise Exception(error)

        await manager.send_progress({
            "type": "progress",
            "step": "evaluation",
            "message": f"Evaluating {len(options)} strategic options in parallel..."
        })

        # Evaluate all options IN PARALLEL for performance
        async def evaluate_option(option, index):
            """Evaluate a single strategic option"""
            eval_context = {
                'option': option,
                'epic_context': epic_context
            }

            evaluation, eval_error = await asyncio.to_thread(
                evaluator.run,
                eval_context
            )

            if eval_error:
                print(f"Warning: Failed to evaluate option {index+1}: {eval_error}")
                evaluation = {}
            else:
                print(f"DEBUG: Option {index+1} evaluated successfully")

            return {
                **option,
                "evaluation": evaluation
            }

        # Execute all evaluations in parallel
        print(f"DEBUG: Starting parallel evaluation of {len(options)} options")
        evaluation_tasks = [
            evaluate_option(option, i) for i, option in enumerate(options)
        ]
        evaluated_options = await asyncio.gather(*evaluation_tasks)
        print(f"DEBUG: Parallel evaluation complete")

        await manager.send_progress({
            "type": "progress",
            "step": "evaluation",
            "message": f"Evaluated all {len(options)} options"
        })

        await manager.send_progress({
            "type": "complete",
            "message": "Analysis complete"
        })

        # Cache attachments for later use in Test Ticket Generation
        async with epic_attachments_lock:
            epic_attachments_cache[epic_key] = {
                "epic_attachments": epic_attachments,
                "child_attachments": child_attachments
            }
        print(f"DEBUG: Cached {len(epic_attachments)} epic attachments and {len(child_attachments)} child attachment groups for {epic_key}")

        response_data = {
            "epic_key": epic_key,
            "options": evaluated_options,
            "generated_at": datetime.now().isoformat(),
            "improved_tickets": improved_tickets
        }

        print(f"\n{'='*80}")
        print(f"DEBUG: Returning analysis response for {epic_key}")
        print(f"DEBUG: Number of options: {len(evaluated_options)}")
        print(f"DEBUG: Response structure: {list(response_data.keys())}")
        print(f"{'='*80}\n")

        return response_data

    except Exception as e:
        await manager.send_progress({
            "type": "error",
            "message": str(e)
        })
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/epics/{epic_key}/readiness")
async def assess_epic_readiness(epic_key: str):
    """
    Assess Epic readiness by generating questions and prioritizing them.
    Uses Questioner Agent and Gap Analyzer Agent.
    """
    if not jira_client or not llm_client:
        raise HTTPException(status_code=401, detail="Not authenticated")

    # Validate Jira key format
    epic_key = validate_jira_key(epic_key)

    await manager.send_progress({
        "type": "progress",
        "step": "loading_epic",
        "message": f"Loading Epic {epic_key}..."
    })

    try:
        # Load Epic and children
        epic = jira_client.get_issue(epic_key)
        children_raw = jira_client.get_children_of_epic(epic_key)

        # Process epic description
        epic_description = epic.get('fields', {}).get('description', '')
        if isinstance(epic_description, dict):
            epic_description = adf_to_plaintext(epic_description)
        elif epic_description is None:
            epic_description = ''

        # Clean the description to remove out-of-scope content
        epic_description = clean_jira_text_for_llm(epic_description)

        # Transform children and clean their descriptions
        children = []
        for child in children_raw:
            fields = child.get('fields', {})
            desc = fields.get('description', '')

            if isinstance(desc, dict):
                desc = adf_to_plaintext(desc)
            elif desc is None:
                desc = ''

            # Clean child descriptions
            desc = clean_jira_text_for_llm(desc)

            children.append({
                'key': child.get('key', ''),
                'summary': fields.get('summary', ''),
                'description': desc
            })

        epic_data = {
            'key': epic.get('key'),
            'summary': epic.get('fields', {}).get('summary'),
            'description': epic_description
        }

        # Step 1: Generate questions
        await manager.send_progress({
            "type": "progress",
            "step": "generating_questions",
            "message": "Questioner Agent is analyzing Epic for gaps and ambiguities..."
        })

        questioner = QuestionerAgent(llm_client)
        questions, questions_error = await asyncio.to_thread(
            questioner.generate_questions,
            epic_data,
            children
        )

        if questions_error:
            raise Exception(f"Failed to generate questions: {questions_error}")

        # Step 2: Analyze and prioritize questions
        await manager.send_progress({
            "type": "progress",
            "step": "analyzing_gaps",
            "message": "Gap Analyzer is prioritizing questions..."
        })

        analyzer = GapAnalyzerAgent(llm_client)
        analysis, analysis_error = await asyncio.to_thread(
            analyzer.analyze_questions,
            questions,
            epic_data,
            children
        )

        if analysis_error:
            raise Exception(f"Failed to analyze questions: {analysis_error}")

        await manager.send_progress({
            "type": "complete",
            "message": "Readiness assessment complete"
        })

        return {
            "epic_key": epic_key,
            "readiness_assessment": analysis,
            "generated_at": datetime.now().isoformat()
        }

    except Exception as e:
        await manager.send_progress({
            "type": "error",
            "message": str(e)
        })
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/tickets/improve")
async def improve_ticket(request: dict):
    """
    Generate an improved version of a ticket using Ticket Improver Agent.
    """
    if not llm_client:
        raise HTTPException(status_code=401, detail="Not authenticated")

    await manager.send_progress({
        "type": "progress",
        "step": "improving_ticket",
        "message": "Ticket Improver Agent is analyzing and enhancing the ticket..."
    })

    try:
        ticket_data = request.get('ticket')
        questions = request.get('questions', [])
        epic_context = request.get('epic_context')

        if not ticket_data:
            raise HTTPException(status_code=400, detail="Ticket data is required")

        # IMPORTANT: Do NOT clean ticket description for ticket improver
        # The improver needs original out-of-scope content to preserve it
        # Text cleaning is only for test generation agents (to avoid creating tests for removed features)
        # if ticket_data.get('description'):
        #     ticket_data['description'] = clean_jira_text_for_llm(ticket_data['description'])

        # Clean epic context if provided
        if epic_context:
            if epic_context.get('epic_desc'):
                epic_context['epic_desc'] = clean_jira_text_for_llm(epic_context['epic_desc'])
            if epic_context.get('description'):
                epic_context['description'] = clean_jira_text_for_llm(epic_context['description'])
            # Clean child ticket descriptions if they exist
            if epic_context.get('children'):
                for child in epic_context['children']:
                    if child.get('desc'):
                        child['desc'] = clean_jira_text_for_llm(child['desc'])
                    if child.get('description'):
                        child['description'] = clean_jira_text_for_llm(child['description'])

        # Initialize Ticket Improver Agent
        improver = TicketImproverAgent(llm_client)

        # Generate improved version
        improvement, error = await asyncio.to_thread(
            improver.improve_ticket,
            ticket_data,
            questions,
            epic_context
        )

        if error:
            raise Exception(f"Failed to improve ticket: {error}")

        await manager.send_progress({
            "type": "complete",
            "message": "Ticket improvement complete"
        })

        return improvement

    except Exception as e:
        await manager.send_progress({
            "type": "error",
            "message": str(e)
        })
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/tickets/{ticket_key}/improve")
async def improve_ticket_by_key(ticket_key: str):
    """
    Fetch a ticket by key and generate an improved version.
    Standalone endpoint for the Ticket Improver page.
    """
    if not jira_client or not llm_client:
        raise HTTPException(status_code=401, detail="Not authenticated")

    # Validate Jira key format
    ticket_key = validate_jira_key(ticket_key)

    try:
        # Fetch the ticket from Jira
        ticket_data = jira_client.get_issue(ticket_key)
        if not ticket_data:
            raise HTTPException(status_code=404, detail=f"Ticket {ticket_key} not found")

        # Extract fields
        fields = ticket_data.get("fields", {})

        # Process ticket description (convert from ADF if needed)
        ticket_description = fields.get('description', '')
        if isinstance(ticket_description, dict):
            from ai_tester.utils.utils import adf_to_plaintext
            ticket_description = adf_to_plaintext(ticket_description)
        elif ticket_description is None:
            ticket_description = ''

        # Extract acceptance criteria from custom fields
        acceptance_criteria = ""

        # First try the known custom field ID for this Jira instance
        ac_field_value = fields.get("customfield_10524")
        if ac_field_value:
            if isinstance(ac_field_value, str):
                acceptance_criteria = ac_field_value
            elif isinstance(ac_field_value, dict):
                from ai_tester.utils.utils import adf_to_plaintext
                acceptance_criteria = adf_to_plaintext(ac_field_value)

        # Fallback: Look for acceptance criteria in custom fields by name
        if not acceptance_criteria:
            for field_key, field_value in fields.items():
                if "acceptance" in field_key.lower() or "criteria" in field_key.lower():
                    if field_value:
                        if isinstance(field_value, str):
                            acceptance_criteria = field_value
                        elif isinstance(field_value, dict):
                            from ai_tester.utils.utils import adf_to_plaintext
                            acceptance_criteria = adf_to_plaintext(field_value)
                        break

        # IMPORTANT: Do NOT clean ticket description for ticket improver
        # The improver needs original out-of-scope content to preserve it
        # Text cleaning is only for test generation agents (to avoid creating tests for removed features)
        # if ticket_description:
        #     ticket_description = clean_jira_text_for_llm(ticket_description)

        # Build ticket dict for improver
        ticket_dict = {
            'key': ticket_data.get('key'),
            'summary': fields.get('summary', ''),
            'description': ticket_description,
            'acceptance_criteria': acceptance_criteria
        }

        # Initialize Ticket Improver Agent
        improver = TicketImproverAgent(llm_client)

        # Generate improved version (no questions or epic context for standalone use)
        improvement, error = await asyncio.to_thread(
            improver.improve_ticket,
            ticket_dict,
            None,  # No questions
            None   # No epic context
        )

        if error:
            raise Exception(f"Failed to improve ticket: {error}")

        return {
            "success": True,
            "improved_ticket": improvement.get('improved_ticket'),
            "improvements_made": improvement.get('improvements_made', []),
            "quality_increase": improvement.get('quality_increase', 0)
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/test-tickets/generate")
async def generate_test_tickets(request: TestTicketGenerationRequest):
    """
    Generate test tickets based on selected strategic option.
    """
    if not jira_client or not llm_client:
        raise HTTPException(status_code=401, detail="Not authenticated")

    try:
        print(f"\n{'='*80}")
        print(f"DEBUG: Starting test ticket generation for epic {request.epic_key}")
        print(f"DEBUG: Selected option index: {request.selected_option_index}")
        print(f"DEBUG: Option data provided: {request.selected_option is not None}")
        print(f"{'='*80}\n")

        # Load Epic and children
        await manager.send_progress({
            "type": "progress",
            "step": "loading_epic",
            "message": f"Loading Epic {request.epic_key}..."
        })

        epic = jira_client.get_issue(request.epic_key)
        children_raw = jira_client.get_children_of_epic(request.epic_key)

        # Process epic description
        epic_description = epic.get('fields', {}).get('description', '')
        if isinstance(epic_description, dict):
            epic_description = adf_to_plaintext(epic_description)
        elif epic_description is None:
            epic_description = ''

        # Clean epic description FOR TEST GENERATION
        # NOTE: We keep the original description for coverage review (passed separately)
        # so that the "Out of Scope" section is visible to the coverage reviewer
        epic_description_for_test_gen = clean_jira_text_for_llm(epic_description)

        # Transform children to expected format and clean descriptions
        children = []
        for child in children_raw:
            fields = child.get('fields', {})
            desc = fields.get('description', '')

            # Convert ADF description to plaintext if needed
            if isinstance(desc, dict):
                desc = adf_to_plaintext(desc)
            elif desc is None:
                desc = ''

            # Clean description
            desc = clean_jira_text_for_llm(desc)

            children.append({
                'key': child.get('key', ''),
                'summary': fields.get('summary', ''),
                'desc': desc
            })

        # Load and process attachments
        await manager.send_progress({
            "type": "progress",
            "step": "loading_attachments",
            "message": "Processing epic attachments..."
        })

        # Check cache first for attachments (includes uploaded documents from Epic Analysis)
        cache_hit = False
        async with epic_attachments_lock:
            if request.epic_key in epic_attachments_cache:
                print(f"DEBUG: Using cached attachments for {request.epic_key}")
                cached_data = epic_attachments_cache[request.epic_key]
                epic_attachments = cached_data.get("epic_attachments", [])
                child_attachments = cached_data.get("child_attachments", {})
                print(f"DEBUG: Retrieved {len(epic_attachments)} cached epic attachments and {len(child_attachments)} child attachment groups")
                cache_hit = True

        if not cache_hit:
            print(f"DEBUG: No cached attachments found for {request.epic_key}, fetching from Jira")
            epic_attachments = []
            child_attachments = {}

            # Process epic attachments
            epic_attachment_list = jira_client.get_attachments(request.epic_key)
            for attachment in epic_attachment_list:
                processed = jira_client.process_attachment(attachment)
                if processed:
                    epic_attachments.append(processed)

            print(f"DEBUG: Processed {len(epic_attachments)} attachments for epic {request.epic_key}")

            # Also extract attachments referenced in the description (embedded media)
            print(f"DEBUG: Checking for attachments embedded in description (test ticket generation)")
            epic_desc_field = epic.get('fields', {}).get('description')
            if epic_desc_field and isinstance(epic_desc_field, dict):
                embedded_attachments = jira_client.extract_attachments_from_description(request.epic_key, epic_desc_field)
                print(f"DEBUG: Found {len(embedded_attachments)} attachments embedded in description")

                # Process any embedded attachments that aren't already in the list
                already_processed = {att.get('filename') for att in epic_attachments}
                for attachment in embedded_attachments:
                    filename = attachment.get('filename')
                    if filename not in already_processed:
                        print(f"DEBUG: Processing embedded attachment: {filename}")
                        processed = jira_client.process_attachment(attachment)
                        if processed:
                            epic_attachments.append(processed)
                            print(f"DEBUG: Successfully processed embedded: {processed.get('filename')}")
                    else:
                        print(f"DEBUG: Embedded attachment {filename} already processed")

            print(f"DEBUG: Total epic attachments after checking description: {len(epic_attachments)}")

            # Process child ticket attachments (limit to avoid excessive API calls)
            # Only do this if we didn't get cached attachments
            for idx, child in enumerate(children_raw[:20]):  # Only first 20 children to avoid slowdown
                child_key = child.get('key')
                if child_key:
                    child_attachment_list = jira_client.get_attachments(child_key)
                    processed_child_attachments = []
                    for attachment in child_attachment_list:
                        processed = jira_client.process_attachment(attachment)
                        if processed:
                            processed_child_attachments.append(processed)

                    if processed_child_attachments:
                        child_attachments[child_key] = processed_child_attachments

        # Process uploaded documents
        if 'files' in locals() and files and len(files) > 0:
            if 'manager' in locals():
                await manager.send_progress({
                    "type": "progress",
                    "step": "processing_uploads",
                    "message": f"Processing {len(files)} uploaded documents..."
                })

            print(f"DEBUG: Processing {len(files)} uploaded documents")

            for uploaded_file in files:
                try:
                    file_bytes = await uploaded_file.read()
                    print(f"DEBUG: Processing uploaded file: {uploaded_file.filename} ({uploaded_file.content_type}, {len(file_bytes)} bytes)")

                    result = {
                        "filename": uploaded_file.filename,
                        "mime_type": uploaded_file.content_type,
                        "size": len(file_bytes)
                    }

                    # PDF files
                    if uploaded_file.content_type == "application/pdf" or uploaded_file.filename.lower().endswith('.pdf'):
                        from ai_tester.utils.utils import extract_text_from_pdf
                        text = extract_text_from_pdf(file_bytes)
                        result["type"] = "document"
                        result["content"] = text
                        epic_attachments.append(result)
                        print(f"DEBUG: Successfully processed uploaded PDF: {uploaded_file.filename}")

                    # Word documents
                    elif uploaded_file.content_type in ["application/vnd.openxmlformats-officedocument.wordprocessingml.document", "application/msword"] or uploaded_file.filename.lower().endswith(('.docx', '.doc')):
                        from ai_tester.utils.utils import extract_text_from_word
                        text = extract_text_from_word(file_bytes)
                        result["type"] = "document"
                        result["content"] = text
                        epic_attachments.append(result)
                        print(f"DEBUG: Successfully processed uploaded Word doc: {uploaded_file.filename}")

                    # Images
                    elif uploaded_file.content_type and uploaded_file.content_type.startswith("image/"):
                        if uploaded_file.content_type in ["image/jpeg", "image/png", "image/gif", "image/webp"]:
                            from ai_tester.utils.utils import encode_image_to_base64
                            base64_data = encode_image_to_base64(file_bytes, uploaded_file.content_type)
                            result["type"] = "image"
                            result["content"] = base64_data
                            result["data_url"] = f"data:{uploaded_file.content_type};base64,{base64_data}"
                            epic_attachments.append(result)
                            print(f"DEBUG: Successfully processed uploaded image: {uploaded_file.filename}")

                    # Text/Markdown files
                    elif uploaded_file.content_type and (uploaded_file.content_type.startswith("text/") or uploaded_file.filename.lower().endswith(('.txt', '.md'))):
                        try:
                            text = file_bytes.decode('utf-8')
                            result["type"] = "document"
                            result["content"] = text
                            epic_attachments.append(result)
                            print(f"DEBUG: Successfully processed uploaded text file: {uploaded_file.filename}")
                        except Exception as e:
                            print(f"DEBUG: Error decoding text file {uploaded_file.filename}: {e}")

                except Exception as e:
                    print(f"DEBUG: Error processing uploaded file {uploaded_file.filename}: {e}")

            print(f"DEBUG: Total epic attachments after processing uploads: {len(epic_attachments)}")

        epic_context = {
            'epic_key': epic.get('key'),
            'epic_summary': epic.get('fields', {}).get('summary'),
            'epic_desc': epic_description_for_test_gen,  # Cleaned version for test generation
            'children': children,
            'epic_attachments': epic_attachments,
            'child_attachments': child_attachments
        }

        # Use provided option data if available, otherwise re-analyze (legacy behavior)
        if request.selected_option:
            print("DEBUG: Using provided strategic option (no re-analysis needed)")
            selected_option = request.selected_option
        else:
            print("DEBUG: No option data provided, re-analyzing epic...")
            await manager.send_progress({
                "type": "progress",
                "step": "loading_analysis",
                "message": "Analyzing epic strategy..."
            })

            # Re-run analysis (legacy behavior for backwards compatibility)
            planner = StrategicPlannerAgent(llm_client)
            evaluator = EvaluationAgent(llm_client)

            options, error = await asyncio.to_thread(planner.run, epic_context)
            if error:
                raise Exception(error)

            # Check if options list is empty
            if not options or len(options) == 0:
                raise Exception("Strategic planner returned no options. Please try analyzing the epic again.")

            # Get selected option or use best scored
            if request.selected_option_index is not None and request.selected_option_index < len(options):
                selected_option = options[request.selected_option_index]
            else:
                # Auto-select best option if not specified
                for option in options:
                    eval_context = {'option': option, 'epic_context': epic_context}
                    evaluation, eval_error = await asyncio.to_thread(evaluator.run, eval_context)
                    if not eval_error:
                        option['evaluation'] = evaluation
                # Select highest scored
                selected_option = max(options, key=lambda x: x.get('evaluation', {}).get('overall', 0))

        # Generate test tickets for each ticket in the selected strategy
        await manager.send_progress({
            "type": "progress",
            "step": "generating",
            "message": f"Generating {len(selected_option.get('test_tickets', []))} test tickets..."
        })

        generator = TestTicketGeneratorAgent(llm_client)
        reviewer = TestTicketReviewerAgent(llm_client)

        generated_tickets = []
        test_ticket_plans = selected_option.get('test_tickets', [])

        for idx, ticket_plan in enumerate(test_ticket_plans, 1):
            await manager.send_progress({
                "type": "progress",
                "step": "generating",
                "message": f"Generating ticket {idx} of {len(test_ticket_plans)}: {ticket_plan.get('title', 'Test Ticket')}..."
            })

            functional_area = ticket_plan.get('title', f'Area {idx}')

            # Determine which child tickets this test ticket covers
            scope_text = ticket_plan.get('scope', '')
            covered_children = []
            for child in children:
                child_key = child.get('key', '')
                if child_key and child_key in scope_text:
                    covered_children.append({
                        'key': child_key,
                        'summary': child.get('summary', '')
                    })

            # Generate ticket
            gen_context = {
                'epic_name': epic_context['epic_summary'],
                'functional_area': functional_area,
                'child_tickets': covered_children or children[:5],  # Use all if none specified
                'epic_context': epic_context
            }

            ticket_data, gen_error = await asyncio.to_thread(generator.run, gen_context)

            if gen_error:
                print(f"Error generating ticket {idx}: {gen_error}")
                continue

            # Review the ticket
            await manager.send_progress({
                "type": "progress",
                "step": "reviewing",
                "message": f"Reviewing ticket {idx}..."
            })

            review_context = {
                'ticket_data': ticket_data,
                'epic_context': epic_context
            }
            review_data, review_error = await asyncio.to_thread(reviewer.run, review_context)

            if review_error:
                print(f"Error reviewing ticket {idx}: {review_error}")
                review_data = {'quality_score': 50, 'needs_improvement': True}

            # If quality is low, try one refinement
            if review_data.get('quality_score', 0) < 70 and review_data.get('needs_improvement'):
                await manager.send_progress({
                    "type": "progress",
                    "step": "refining",
                    "message": f"Refining ticket {idx} based on feedback..."
                })

                refine_context = {
                    **gen_context,
                    'previous_attempt': json.dumps(ticket_data),
                    'reviewer_feedback': review_data
                }
                ticket_data, refine_error = await asyncio.to_thread(generator.run, refine_context)

                if not refine_error:
                    # Review again
                    review_context['ticket_data'] = ticket_data
                    review_data, _ = await asyncio.to_thread(reviewer.run, review_context)

            # Create TestTicket object
            ticket_id = f"{request.epic_key}-TT-{idx:03d}"

            # Use child_tickets from LLM response if available, otherwise use our inference
            llm_child_tickets = ticket_data.get('child_tickets', [])
            final_child_tickets = llm_child_tickets if llm_child_tickets else covered_children

            test_ticket = TestTicket(
                id=ticket_id,
                summary=ticket_data.get('summary', f'{functional_area}'),
                description=ticket_data.get('description', ''),
                acceptance_criteria=ticket_data.get('acceptance_criteria', []),
                quality_score=review_data.get('quality_score'),
                review_feedback=review_data,
                raw_response=json.dumps(ticket_data),
                epic_key=request.epic_key,
                child_tickets=final_child_tickets,
                functional_area=functional_area,
                selected_option_index=request.selected_option_index,
                strategic_option=selected_option,
                analyzed=False
            )

            # Store ticket
            async with test_tickets_lock:
                test_tickets_storage[ticket_id] = test_ticket
            generated_tickets.append(test_ticket.to_dict())

        # Perform coverage review
        await manager.send_progress({
            "type": "progress",
            "step": "coverage_review",
            "message": "Coverage Reviewer is analyzing test ticket completeness..."
        })

        coverage_reviewer = CoverageReviewerAgent(llm_client)

        epic_data_for_review = {
            'key': epic.get('key'),
            'summary': epic.get('fields', {}).get('summary'),
            'description': epic_description
        }

        # Separate existing test tickets from functional tickets
        def is_test_ticket(child):
            """Determine if a child ticket is a test ticket based on keywords"""
            fields = child.get('fields', {})
            summary = fields.get('summary', '').lower()
            desc = fields.get('description', '')
            if isinstance(desc, dict):
                desc = adf_to_plaintext(desc).lower()
            elif desc is None:
                desc = ''
            else:
                desc = desc.lower()

            # Keywords that indicate a test ticket
            test_keywords = [
                'test:', 'test case', 'testing', 'e2e test', 'end-to-end test',
                'integration test', 'unit test', 'qa:', 'verify', 'validation',
                'acceptance test', 'regression test', 'smoke test', 'sanity test'
            ]

            full_text = f"{summary} {desc}"
            return any(keyword in full_text for keyword in test_keywords)

        existing_test_tickets = []
        functional_tickets = []

        for child in children_raw:
            fields = child.get('fields', {})
            desc = fields.get('description', '')
            if isinstance(desc, dict):
                desc = adf_to_plaintext(desc)
            elif desc is None:
                desc = ''
            desc = clean_jira_text_for_llm(desc)

            ticket_info = {
                'key': child.get('key', ''),
                'summary': fields.get('summary', ''),
                'description': desc
            }

            if is_test_ticket(child):
                existing_test_tickets.append(ticket_info)
            else:
                functional_tickets.append(ticket_info)

        # Combine generated tickets with existing test tickets for coverage analysis
        all_test_tickets = existing_test_tickets + generated_tickets

        print(f"DEBUG: About to call coverage_reviewer.review_coverage with {len(epic_attachments)} epic attachments and {len(child_attachments)} child attachment groups")
        coverage_review, review_error = await asyncio.to_thread(
            coverage_reviewer.review_coverage,
            epic_data_for_review,
            functional_tickets,  # Only functional tickets need to be covered
            all_test_tickets,  # All test tickets (existing + generated)
            epic_attachments,  # Epic attachments for additional context
            child_attachments  # Child ticket attachments for additional context
        )

        if review_error:
            print(f"Warning: Coverage review failed: {review_error}")
            coverage_review = None

        await manager.send_progress({
            "type": "complete",
            "message": f"Successfully generated {len(generated_tickets)} test tickets ({len(existing_test_tickets)} existing test tickets found)"
        })

        return {
            "success": True,
            "test_tickets": generated_tickets,
            "count": len(generated_tickets),
            "epic_key": request.epic_key,
            "coverage_review": coverage_review,
            "epic_data": epic_data_for_review,
            "child_tickets": functional_tickets,
            "existing_test_tickets": existing_test_tickets,
            "existing_test_count": len(existing_test_tickets)
        }

    except Exception as e:
        print(f"\n{'='*80}")
        print(f"ERROR: Test ticket generation failed!")
        print(f"ERROR: Exception type: {type(e).__name__}")
        print(f"ERROR: Exception message: {str(e)}")

        # Print full traceback for debugging
        import traceback
        print("ERROR: Full traceback:")
        print(traceback.format_exc())
        print(f"{'='*80}\n")

        await manager.send_progress({
            "type": "error",
            "message": str(e)
        })
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/test-tickets/fix-coverage")
async def fix_coverage_gaps(request: dict):
    """
    Generate fixes for coverage gaps using Requirements Fixer Agent.
    """
    if not llm_client:
        raise HTTPException(status_code=401, detail="Not authenticated")

    await manager.send_progress({
        "type": "progress",
        "step": "fixing_coverage",
        "message": "Requirements Fixer is generating solutions for coverage gaps..."
    })

    try:
        coverage_review = request.get('coverage_review')
        existing_tickets = request.get('existing_tickets', [])
        epic_data = request.get('epic_data')
        child_tickets = request.get('child_tickets', [])

        if not coverage_review:
            raise HTTPException(status_code=400, detail="Coverage review is required")

        # Initialize Requirements Fixer Agent
        fixer = RequirementsFixerAgent(llm_client)

        # Generate fixes
        fixes, error = await asyncio.to_thread(
            fixer.generate_fixes,
            coverage_review,
            existing_tickets,
            epic_data,
            child_tickets
        )

        if error:
            raise Exception(f"Failed to generate fixes: {error}")

        await manager.send_progress({
            "type": "complete",
            "message": "Coverage fixes generated successfully"
        })

        return fixes

    except Exception as e:
        await manager.send_progress({
            "type": "error",
            "message": str(e)
        })
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/test-tickets/apply-fixes")
async def apply_coverage_fixes(request: dict):
    """
    Apply coverage fixes to test tickets and recalculate coverage.

    Creates new tickets, updates existing ones, then re-runs coverage review
    to provide updated coverage percentage.
    """
    if not jira_client or not llm_client:
        raise HTTPException(status_code=401, detail="Not authenticated")

    try:
        epic_key = request.get('epic_key')
        new_tickets = request.get('new_tickets', [])
        ticket_updates = request.get('ticket_updates', [])
        epic_data = request.get('epic_data')
        child_tickets = request.get('child_tickets', [])

        if not epic_key:
            raise HTTPException(status_code=400, detail="Epic key is required")

        applied_tickets = []

        # Create new test tickets
        for idx, new_ticket in enumerate(new_tickets):
            ticket_id = f"{epic_key}-FIX-{idx+1:03d}"

            # Get child_tickets in the new format [{key, summary}]
            # Support both old format (covers_child_tickets: ["KEY"]) and new format (child_tickets: [{key, summary}])
            child_tickets_data = new_ticket.get('child_tickets', [])
            if not child_tickets_data and 'covers_child_tickets' in new_ticket:
                # Convert old format to new format by looking up summaries from child_tickets list
                old_format = new_ticket.get('covers_child_tickets', [])
                child_tickets_data = []
                for ticket_key in old_format:
                    # Find the child ticket details from the input child_tickets
                    ticket_info = next((t for t in child_tickets if t.get('key') == ticket_key), None)
                    if ticket_info:
                        child_tickets_data.append({
                            "key": ticket_key,
                            "summary": ticket_info.get('summary', '')
                        })
                    else:
                        child_tickets_data.append({
                            "key": ticket_key,
                            "summary": ""
                        })

            test_ticket = TestTicket(
                id=ticket_id,
                summary=new_ticket.get('summary', ''),
                description=new_ticket.get('description', ''),
                acceptance_criteria=new_ticket.get('acceptance_criteria', []),
                quality_score=85,  # Assume good quality since generated by fixer
                epic_key=epic_key,
                child_tickets=child_tickets_data,
                functional_area=new_ticket.get('addresses_gap', 'Coverage Fix'),
                analyzed=False
            )

            # Store ticket
            async with test_tickets_lock:
                test_tickets_storage[ticket_id] = test_ticket
            applied_tickets.append(test_ticket.to_dict())

        # Update existing tickets
        for update in ticket_updates:
            ticket_id = update.get('original_ticket_id')
            async with test_tickets_lock:
                if ticket_id in test_tickets_storage:
                    ticket = test_tickets_storage[ticket_id]
                    ticket.summary = update.get('updated_summary', ticket.summary)
                    ticket.description = update.get('updated_description', ticket.description)
                    ticket.acceptance_criteria = update.get('updated_acceptance_criteria', ticket.acceptance_criteria)
                    applied_tickets.append(ticket.to_dict())

        # Re-run coverage review with updated test tickets
        updated_coverage_review = None
        if epic_data and child_tickets:
            try:
                # Get all test tickets for this epic
                all_test_tickets = [
                    ticket.to_dict()
                    for ticket in test_tickets_storage.values()
                    if ticket.epic_key == epic_key
                ]

                # Fetch attachments for coverage review context
                epic_attachments = []
                child_attachments = {}

                # Process epic attachments
                epic_attachment_list = jira_client.get_attachments(epic_key)
                for attachment in epic_attachment_list:
                    processed = jira_client.process_attachment(attachment)
                    if processed:
                        epic_attachments.append(processed)

                # Process child ticket attachments
                for child in child_tickets[:20]:  # Limit to first 20 to avoid slowdown
                    child_key = child.get('key')
                    if child_key:
                        child_attachment_list = jira_client.get_attachments(child_key)
                        processed_child_attachments = []
                        for attachment in child_attachment_list:
                            processed = jira_client.process_attachment(attachment)
                            if processed:
                                processed_child_attachments.append(processed)

                        if processed_child_attachments:
                            child_attachments[child_key] = processed_child_attachments

                # Re-run coverage review
                coverage_reviewer = CoverageReviewerAgent(llm_client)
                updated_coverage_review, error = await asyncio.to_thread(
                    coverage_reviewer.review_coverage,
                    epic_data,
                    child_tickets,
                    all_test_tickets,
                    epic_attachments,
                    child_attachments
                )

                if error:
                    print(f"Warning: Failed to recalculate coverage: {error}")
            except Exception as e:
                print(f"Warning: Failed to recalculate coverage: {e}")

        return {
            "success": True,
            "applied_count": len(applied_tickets),
            "tickets": applied_tickets,
            "updated_coverage_review": updated_coverage_review
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ============================================================================
# Test Case Generation
# ============================================================================

def transform_improved_ticket_for_frontend(improved_ticket_data):
    """
    Transform improved ticket data from backend format to frontend format.

    Backend provides:
    - acceptance_criteria_grouped: List[AcceptanceCriteriaCategory] (each category has name and criteria list)
    - testing_notes: str
    - edge_cases: List[str]

    Frontend expects:
    - acceptance_criteria: List[str] (flat list)
    - technical_notes: str (renamed from testing_notes)
    - edge_cases: List[str]
    - error_scenarios: List[str] (optional)
    """
    if not improved_ticket_data:
        return None

    transformed = {
        "summary": improved_ticket_data.get("summary", ""),
        "description": improved_ticket_data.get("description", ""),
        "edge_cases": improved_ticket_data.get("edge_cases", []),
        "error_scenarios": improved_ticket_data.get("error_scenarios", [])
    }

    # Flatten acceptance_criteria_grouped to acceptance_criteria
    ac_grouped = improved_ticket_data.get("acceptance_criteria_grouped", [])
    if ac_grouped:
        flat_ac = []
        for category in ac_grouped:
            # Each category is a dict with 'name' and 'criteria' keys
            if isinstance(category, dict):
                criteria_list = category.get("criteria", [])
                flat_ac.extend(criteria_list)
        transformed["acceptance_criteria"] = flat_ac
    else:
        transformed["acceptance_criteria"] = []

    # Rename testing_notes to technical_notes
    transformed["technical_notes"] = improved_ticket_data.get("testing_notes", "")

    return transformed


@app.post("/api/test-cases/generate")
async def generate_test_cases(request: TestCaseGenerationRequest):
    """Generate test cases for a Jira ticket using multi-agent workflow."""
    if not jira_client or not llm_client:
        raise HTTPException(status_code=401, detail="Not authenticated")

    await manager.send_progress({
        "type": "progress",
        "step": "loading_ticket",
        "message": f"Loading ticket {request.ticket_key}..."
    })

    try:
        # Load ticket
        ticket = jira_client.get_issue(request.ticket_key)
        if not ticket:
            raise HTTPException(status_code=404, detail=f"Ticket {request.ticket_key} not found")

        fields = ticket.get("fields", {})
        summary = fields.get("summary", "")

        # Get description
        description = fields.get("description", "")
        if isinstance(description, dict):
            from ai_tester.utils.utils import adf_to_plaintext
            description = adf_to_plaintext(description)

        # Clean description to remove strikethrough and out-of-scope content
        if description:
            description = clean_jira_text_for_llm(description)

        # Get acceptance criteria from custom field OR extract from description
        acceptance_criteria = ""
        for field_key, field_value in fields.items():
            if "acceptance" in field_key.lower() or "criteria" in field_key.lower():
                if isinstance(field_value, str):
                    acceptance_criteria = field_value
                elif isinstance(field_value, dict):
                    from ai_tester.utils.utils import adf_to_plaintext
                    acceptance_criteria = adf_to_plaintext(field_value)

        # If no separate AC field found, try extracting from description
        if not acceptance_criteria and description:
            # Use the analyzer's extraction logic
            analyzer = TicketAnalyzerAgent(llm_client, jira_client)
            ac_blocks = analyzer._extract_acceptance_criteria(description)
            if ac_blocks:
                acceptance_criteria = "\n".join(ac_blocks)

        # Clean acceptance criteria to remove strikethrough and out-of-scope content
        if acceptance_criteria:
            acceptance_criteria = clean_jira_text_for_llm(acceptance_criteria)

        # Step 2: Preprocess ticket with TicketImproverAgent for consistency
        await manager.send_progress({
            "type": "progress",
            "step": "preprocessing",
            "message": "Preprocessing ticket for better analysis..."
        })

        improved_ticket_data = None
        analysis_description = description
        analysis_ac = acceptance_criteria

        try:
            # Check cache first for improved ticket
            async with improved_tickets_lock:
                if request.ticket_key in improved_tickets_cache:
                    print(f"DEBUG: Using cached improved ticket for {request.ticket_key}")
                    improvement_result = improved_tickets_cache[request.ticket_key]
                    improved_ticket_data = improvement_result.get("improved_ticket", {})

            if improved_ticket_data is None:
                # Not in cache, call LLM
                from ai_tester.agents.ticket_improver_agent import TicketImproverAgent
                ticket_improver = TicketImproverAgent(llm_client)

                # Preprocess the ticket
                improvement_result, improve_error = await asyncio.to_thread(
                    ticket_improver.improve_ticket,
                    {
                        "key": request.ticket_key,
                        "summary": summary,
                        "description": description
                    },
                    None,  # questions
                    None,  # epic_context
                    'gpt-4o-mini'  # Use cheaper model for data extraction
                )

                if improvement_result and not improve_error:
                    # Cache the result
                    async with improved_tickets_lock:
                        improved_tickets_cache[request.ticket_key] = improvement_result
                    improved_ticket_data = improvement_result.get("improved_ticket", {})
                    print(f"DEBUG: Ticket preprocessed and cached for {request.ticket_key}. Quality increase: {improvement_result.get('quality_increase', 0)}%")
                else:
                    print(f"DEBUG: Ticket preprocessing failed: {improve_error}. Using original ticket.")

            # Use improved description and AC for test case generation
            if improved_ticket_data:
                if improved_ticket_data.get("description"):
                    analysis_description = improved_ticket_data["description"]

                if improved_ticket_data.get("acceptance_criteria"):
                    ac_list = improved_ticket_data["acceptance_criteria"]
                    if isinstance(ac_list, list):
                        analysis_ac = "\n".join([f"- {ac}" for ac in ac_list])
                    else:
                        analysis_ac = str(ac_list)
        except Exception as e:
            print(f"DEBUG: Ticket preprocessing error: {e}. Using original ticket.")

        await manager.send_progress({
            "type": "progress",
            "step": "generating",
            "substep": "generation",
            "message": "AI is generating test cases (this may take 30-60 seconds)..."
        })

        # Import helper functions from generate_test_cases.py
        from ai_tester.utils.test_case_generator import critic_review, fixer, generate_test_cases_with_retry

        # Get the current event loop for thread-safe WebSocket updates
        loop = asyncio.get_event_loop()

        # Define progress callback to send WebSocket updates
        # This is called from a thread, so we use run_coroutine_threadsafe
        def send_progress_update(substep: str, message: str):
            try:
                asyncio.run_coroutine_threadsafe(
                    manager.send_progress({
                        "type": "progress",
                        "step": "generating",
                        "substep": substep,
                        "message": message
                    }),
                    loop
                )
            except Exception as e:
                print(f"Warning: Failed to send progress update: {e}")

        # Build prompts
        sys_prompt = """You are an expert QA test case designer. Your task is to analyze Jira tickets and create comprehensive, detailed test cases.

CRITICAL: BLACK-BOX MANUAL TESTING FOCUS
- These test cases are for MANUAL TESTERS performing BLACK-BOX testing
- Focus on USER-OBSERVABLE behavior and outcomes
- NO technical implementation details (no code, APIs, databases, or internal logic)
- Write from the perspective of what a user can see, click, and verify
- Steps should be executable by someone without technical knowledge of the system internals

REQUIREMENT EXTRACTION GUIDELINES:
A "requirement" is a single, testable capability or constraint that can be verified independently.

CORRECT Requirement Granularity (aim for 3-7 requirements per ticket):
✓ "User can submit an enquiry form with all required fields"
✓ "System validates email format on submission"
✓ "Form displays success message after submission"
✓ "User can attach files up to 10MB"

INCORRECT - Too Granular (avoid breaking into 10+ micro-requirements):
✗ "Form has a name field"
✗ "Form has an email field"
✗ "Form has a message field"
✗ "Form has a submit button"
(These should be ONE requirement: "Form displays with all required fields")

INCORRECT - Too Broad (avoid vague mega-requirements):
✗ "Form works correctly"
✗ "User can use the system"
(Too vague - break into specific testable capabilities)

GUIDELINE: Extract 3-7 meaningful requirements per ticket. Group related fields/buttons into single requirements. Focus on distinct user capabilities, not individual UI elements.

TESTING PHILOSOPHY:
For EACH requirement identified, create exactly THREE test cases:
1. One POSITIVE test (happy path)
2. One NEGATIVE test (error handling)
3. One EDGE CASE test (boundary conditions)

This ensures complete coverage with clear traceability.

REQUIRED JSON OUTPUT:
{
  "requirements": [
    {"id": "REQ-001", "description": "Clear, testable requirement statement", "source": "Acceptance Criteria"}
  ],
  "test_cases": [
    {
      "requirement_id": "REQ-001",
      "requirement_desc": "Brief summary",
      "title": "REQ-001 Positive: Title",
      "priority": 1,
      "test_type": "Positive",
      "tags": ["tag1", "tag2"],
      "steps": [
        "Step 1: Specific action",
        "Expected Result: Expected outcome",
        "Step 2: Next action",
        "Expected Result: Next expected outcome"
      ]
    }
  ]
}

MANDATORY RULES:
1. Extract 3-7 meaningful requirements (group related items, avoid micro-requirements)
2. For EACH requirement, create EXACTLY 3 test cases
3. Formula: N requirements → N × 3 test cases (typically 9-21 test cases total)
4. Each test case must have 3+ detailed steps (simple tests need 3 steps, complex scenarios may need 8+ steps - use your judgment based on what the test logically requires)
5. ⚠️ CRITICAL STEP FORMAT: EVERY "Step N:" line MUST be immediately followed by an "Expected Result:" line. This alternating format is MANDATORY and non-negotiable. No exceptions!
6. CONSISTENCY: Always extract the same requirements for the same input - be deterministic"""

        user_prompt = f"""Analyze this Jira ticket and generate comprehensive test cases:

TICKET: {request.ticket_key}
SUMMARY: {summary}

DESCRIPTION:
{analysis_description[:2000]}

ACCEPTANCE CRITERIA:
{analysis_ac if analysis_ac else "No explicit acceptance criteria provided - extract requirements from description"}

Generate test cases following the 3-per-requirement rule (Positive, Negative, Edge Case)."""

        # Generate test cases with critic review
        result, critic_data = await asyncio.to_thread(
            generate_test_cases_with_retry,
            llm=llm_client,
            sys_prompt=sys_prompt,
            user_prompt=user_prompt,
            summary=summary,
            requirements_for_review=None,
            max_retries=2,
            progress_callback=send_progress_update
        )

        if not result:
            raise Exception("Failed to generate test cases after retries")

        requirements = result.get("requirements", [])
        test_cases_raw = result.get("test_cases", [])

        await manager.send_progress({
            "type": "complete",
            "message": f"Generated {len(test_cases_raw)} test cases successfully"
        })

        return TestCaseResponse(
            test_cases=test_cases_raw,
            ticket_info={
                "key": ticket["key"],
                "summary": summary,
                "description": description,
                "acceptance_criteria": acceptance_criteria,
                "requirements_count": len(requirements)
            },
            requirements=requirements,
            generated_at=datetime.now().isoformat(),
            improved_ticket=transform_improved_ticket_for_frontend(improved_ticket_data)
        )

    except Exception as e:
        await manager.send_progress({
            "type": "error",
            "message": str(e)
        })
        raise HTTPException(status_code=500, detail=str(e))


# ============================================================================
# Export Test Cases
# ============================================================================

@app.post("/api/test-cases/export")
async def export_test_cases(request: ExportRequest):
    """Export test cases to CSV, XLSX, or TestRail format."""

    try:
        print(f"DEBUG Export: Received export request - format={request.format}, ticket_key={request.ticket_key}, num_test_cases={len(request.test_cases)}")

        format_type = request.format.lower()
        test_cases = request.test_cases
        ticket_key = request.ticket_key

        if format_type == "xlsx":
            # Generate XLSX file
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = Workbook()
            ws = wb.active
            ws.title = "Test Cases"

            # Headers - Azure DevOps format
            headers = ["ID", "Work Item Type", "Title", "Test Step", "Step Action", "Step Expected"]
            ws.append(headers)

            # Style headers
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Add test cases
            for idx, case in enumerate(test_cases, start=1):
                tc_id = f"TC-{idx:03d}"
                title = case.get("title", "")

                # Prepend TC-ID if not already there
                if not title.startswith("TC-"):
                    title = f"{tc_id} {title}"

                # First row: test case metadata with empty step fields
                # ID column must be EMPTY for new test cases (Azure DevOps requirement)
                ws.append(["", "Test Case", title, "", "", ""])

                # Following rows: steps with actions and expected results
                steps = case.get("steps", [])
                for step in steps:
                    # Handle both string format and object format
                    if isinstance(step, str):
                        # String format: "Step 1: action" and "Expected Result: result"
                        if step.startswith("Step "):
                            step_num = step.split(":")[0].replace("Step ", "").strip()
                            action = step.split(":", 1)[1].strip() if ":" in step else step
                            ws.append(["", "", "", step_num, action, ""])
                        elif step.startswith("Expected Result:"):
                            expected = step.replace("Expected Result:", "").strip()
                            # Update the last row's expected column
                            if ws.max_row > 1:
                                ws.cell(row=ws.max_row, column=6, value=expected)
                    else:
                        # Object format: {"action": "...", "expected_result": "..."}
                        action = step.get("action", step.get("step", ""))
                        expected = step.get("expected_result", step.get("expected", ""))
                        step_num = str(steps.index(step) + 1)
                        ws.append(["", "", "", step_num, action, expected])

            # Auto-size columns (max width 50)
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f"attachment; filename=test_cases_{ticket_key}.xlsx"
                }
            )

        elif format_type == "testrail":
            # Generate TestRail CSV
            output = io.StringIO()
            writer = csv.writer(output)

            # Headers
            writer.writerow(["ID", "Title", "Section", "Priority", "Type", "Steps", "Expected Result"])

            for idx, case in enumerate(test_cases, start=1):
                req_id = case.get("requirement_id", "UNMAPPED")
                tc_id = f"{req_id}:TC-{idx:03d}"
                title = case.get("title", "")
                priority_num = case.get("priority", 2)

                # Convert priority: 1=High, 2=Medium, 3=Low
                priority_text = {1: "High", 2: "Medium", 3: "Low"}.get(priority_num, "Medium")

                test_type = case.get("test_type", "Positive")

                # Format steps as numbered list
                steps = case.get("steps", [])
                steps_list = []
                expected_list = []

                step_counter = 1
                for step in steps:
                    if isinstance(step, str):
                        if step.startswith("Step "):
                            action = step.split(":", 1)[1].strip() if ":" in step else step
                            steps_list.append(f"{step_counter}. {action}")
                            step_counter += 1
                        elif step.startswith("Expected Result:"):
                            expected = step.replace("Expected Result:", "").strip()
                            expected_list.append(f"{len(expected_list) + 1}. {expected}")
                    else:
                        action = step.get("action", step.get("step", ""))
                        expected = step.get("expected_result", step.get("expected", ""))
                        steps_list.append(f"{step_counter}. {action}")
                        expected_list.append(f"{step_counter}. {expected}")
                        step_counter += 1

                steps_text = "\n".join(steps_list)
                expected_text = "\n".join(expected_list)

                writer.writerow([
                    tc_id,
                    title,
                    req_id,
                    priority_text,
                    test_type,
                    steps_text,
                    expected_text
                ])

            output.seek(0)
            return StreamingResponse(
                io.BytesIO(output.getvalue().encode('utf-8')),
                media_type="text/csv",
                headers={
                    "Content-Disposition": f"attachment; filename=test_cases_{ticket_key}_testrail.csv"
                }
            )

        else:  # Generic CSV (Azure DevOps format)
            output = io.StringIO()
            writer = csv.writer(output)

            # Headers
            writer.writerow(["ID", "Work Item Type", "Title", "Test Step", "Step Action", "Step Expected"])

            for idx, case in enumerate(test_cases, start=1):
                tc_id = f"TC-{idx:03d}"
                title = case.get("title", "")

                # Prepend TC-ID if not already there
                if not title.startswith("TC-"):
                    title = f"{tc_id} {title}"

                # First row: test case metadata
                # ID column must be EMPTY for new test cases (Azure DevOps requirement)
                writer.writerow(["", "Test Case", title, "", "", ""])

                # Following rows: steps
                # For string format, collect step-expected pairs first
                steps = case.get("steps", [])
                step_rows = []
                current_action = None
                current_step_num = None
                step_counter = 1

                for step in steps:
                    if isinstance(step, str):
                        if step.startswith("Step "):
                            step_num = step.split(":")[0].replace("Step ", "").strip()
                            action = step.split(":", 1)[1].strip() if ":" in step else step
                            current_action = action
                            current_step_num = step_num
                        elif step.startswith("Expected Result:"):
                            expected = step.replace("Expected Result:", "").strip()
                            # Write the pair now that we have both action and expected
                            if current_action is not None:
                                step_rows.append(["", "", "", current_step_num, current_action, expected])
                                current_action = None
                                current_step_num = None
                    else:
                        action = step.get("action", step.get("step", ""))
                        expected = step.get("expected_result", step.get("expected", ""))
                        step_rows.append(["", "", "", str(step_counter), action, expected])
                        step_counter += 1

                # Write any remaining action without expected result
                if current_action is not None:
                    step_rows.append(["", "", "", current_step_num, current_action, ""])

                # Write all collected rows
                for row in step_rows:
                    writer.writerow(row)

            output.seek(0)
            return StreamingResponse(
                io.BytesIO(output.getvalue().encode('utf-8')),
                media_type="text/csv",
                headers={
                    "Content-Disposition": f"attachment; filename=test_cases_{ticket_key}.csv"
                }
            )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")


# ============================================================================
# Test Ticket Management
# ============================================================================

@app.get("/api/test-tickets")
async def list_test_tickets(epic_key: Optional[str] = None):
    """
    List all generated test tickets, optionally filtered by epic_key.
    """
    async with test_tickets_lock:
        if epic_key:
            filtered_tickets = [
                ticket.to_dict()
                for ticket in test_tickets_storage.values()
                if ticket.epic_key == epic_key
            ]
            return {
                "test_tickets": filtered_tickets,
                "count": len(filtered_tickets),
                "epic_key": epic_key
            }
        else:
            all_tickets = [ticket.to_dict() for ticket in test_tickets_storage.values()]
            return {
                "test_tickets": all_tickets,
                "count": len(all_tickets)
            }


@app.get("/api/test-tickets/{ticket_id}")
async def get_test_ticket(ticket_id: str):
    """
    Get a specific test ticket by ID.
    """
    async with test_tickets_lock:
        if ticket_id not in test_tickets_storage:
            raise HTTPException(status_code=404, detail=f"Test ticket {ticket_id} not found")

        ticket = test_tickets_storage[ticket_id]
        return ticket.to_dict()


@app.post("/api/test-tickets/{ticket_id}/generate-test-cases")
async def generate_test_cases_for_ticket(ticket_id: str):
    """
    Generate test cases for a specific test ticket.
    """
    if not jira_client or not llm_client:
        raise HTTPException(status_code=401, detail="Not authenticated")

    if ticket_id not in test_tickets_storage:
        raise HTTPException(status_code=404, detail=f"Test ticket {ticket_id} not found")

    ticket = test_tickets_storage[ticket_id]

    await manager.send_progress({
        "type": "progress",
        "step": "generating",
        "message": f"Generating test cases for {ticket.summary}..."
    })

    try:
        # Import helper functions from generate_test_cases.py
        from ai_tester.utils.test_case_generator import generate_test_cases_with_retry
        from ai_tester.utils.jira_text_cleaner import clean_jira_text_for_llm

        # Fetch source ticket context if child_tickets are present
        source_tickets_context = ""
        if ticket.child_tickets and len(ticket.child_tickets) > 0:
            await manager.send_progress({
                "type": "progress",
                "step": "generating",
                "substep": "fetching_context",
                "message": f"Fetching context from {len(ticket.child_tickets)} source ticket(s)..."
            })

            source_ticket_details = []
            for child_ticket in ticket.child_tickets:
                try:
                    ticket_key = child_ticket.get("key", "")
                    if ticket_key:
                        # Fetch the actual ticket from Jira
                        source_issue = jira_client.get_issue(ticket_key)
                        source_summary = source_issue.get("fields", {}).get("summary", "")
                        source_description = source_issue.get("fields", {}).get("description", "")

                        # Clean the description
                        if source_description:
                            source_description = clean_jira_text_for_llm(source_description)
                            # Truncate if too long
                            if len(source_description) > 1500:
                                source_description = source_description[:1500] + "..."
                        else:
                            source_description = "No description"

                        # Get acceptance criteria if available
                        source_ac = ""
                        ac_field = source_issue.get("acceptance_criteria", "")
                        if ac_field:
                            source_ac = f"\nAcceptance Criteria:\n{ac_field[:500]}"

                        source_ticket_details.append(f"""
--- Source Ticket: {ticket_key} ---
Summary: {source_summary}
Description:
{source_description}{source_ac}
""")
                except Exception as e:
                    print(f"Warning: Failed to fetch source ticket {child_ticket.get('key', 'unknown')}: {e}")

            if source_ticket_details:
                source_tickets_context = f"""

SOURCE TICKETS CONTEXT:
The test ticket was generated from the following source tickets. Use this context to understand the underlying requirements better:
{"".join(source_ticket_details)}
---
"""

        # Retrieve epic context and attachments if available (for enhanced context)
        # This is optional - standalone mode works without it
        epic_context_text = ""
        if ticket.epic_key:
            try:
                print(f"DEBUG: Fetching epic context for {ticket.epic_key} to enhance test case generation")

                # Fetch epic details
                epic = jira_client.get_issue(ticket.epic_key)
                epic_summary = epic.get('fields', {}).get('summary', '')
                epic_desc_raw = epic.get('fields', {}).get('description', '')
                epic_description = clean_jira_text_for_llm(epic_desc_raw) if epic_desc_raw else "No description"

                # Truncate epic description if too long
                if len(epic_description) > 1000:
                    epic_description = epic_description[:1000] + "..."

                epic_context_text = f"""

EPIC CONTEXT:
This test ticket is part of Epic {ticket.epic_key}:
Epic Summary: {epic_summary}
Epic Description: {epic_description}
"""

                # Check if we have cached attachments for this epic
                async with epic_attachments_lock:
                    if ticket.epic_key in epic_attachments_cache:
                        cached_data = epic_attachments_cache[ticket.epic_key]
                        epic_attachments = cached_data.get("epic_attachments", [])
                    else:
                        epic_attachments = []

                if epic_attachments:
                    print(f"DEBUG: Found {len(epic_attachments)} cached attachments for epic {ticket.epic_key}")
                    epic_context_text += "\nEPIC ATTACHMENTS & DOCUMENTATION:\n"
                    epic_context_text += "The following documents provide additional context:\n"

                    for att in epic_attachments[:5]:  # Limit to 5 most relevant
                        filename = att.get('filename', 'Unknown')
                        att_type = att.get('type', 'unknown')

                        if att_type == 'document':
                            content = att.get('content', '')
                            # Include first 500 chars as preview
                            preview = content[:500] + "..." if len(content) > 500 else content
                            epic_context_text += f"\n  • {filename}:\n    {preview}\n"
                        elif att_type == 'image':
                            epic_context_text += f"\n  • {filename} - UI Mockup/Screenshot (reference for UI testing)\n"

                    epic_context_text += "\nUse this documentation to create more accurate and detailed test cases.\n"

                epic_context_text += "---\n"
                print(f"DEBUG: Epic context prepared ({len(epic_context_text)} characters)")

            except Exception as e:
                print(f"DEBUG: Could not fetch epic context (standalone mode): {e}")
                # Standalone mode - no epic context, which is fine
                epic_context_text = ""

        # Get the current event loop for thread-safe WebSocket updates
        loop = asyncio.get_event_loop()

        def send_progress_update(substep: str, message: str):
            try:
                asyncio.run_coroutine_threadsafe(
                    manager.send_progress({
                        "type": "progress",
                        "step": "generating",
                        "substep": substep,
                        "message": message
                    }),
                    loop
                )
            except Exception as e:
                print(f"Warning: Failed to send progress update: {e}")

        # Build prompts
        sys_prompt = """You are an expert QA test case designer. Your task is to analyze test tickets and create comprehensive, detailed test cases.

CRITICAL: BLACK-BOX MANUAL TESTING FOCUS
- These test cases are for MANUAL TESTERS performing BLACK-BOX testing
- Focus on USER-OBSERVABLE behavior and outcomes
- NO technical implementation details (no code, APIs, databases, or internal logic)
- Write from the perspective of what a user can see, click, and verify
- Steps should be executable by someone without technical knowledge of the system internals

TESTING PHILOSOPHY:
For EACH requirement identified, create exactly THREE test cases:
1. One POSITIVE test (happy path)
2. One NEGATIVE test (error handling)
3. One EDGE CASE test (boundary conditions)

This ensures complete coverage with clear traceability.

REQUIRED JSON OUTPUT:
{
  "requirements": [
    {"id": "REQ-001", "description": "Clear requirement", "source": "Acceptance Criteria"}
  ],
  "test_cases": [
    {
      "requirement_id": "REQ-001",
      "requirement_desc": "Brief summary",
      "title": "REQ-001 Positive: Title",
      "priority": 1,
      "test_type": "Positive",
      "tags": ["tag1", "tag2"],
      "steps": [
        "Step 1: Specific action",
        "Expected Result: Expected outcome",
        "Step 2: Next action",
        "Expected Result: Next expected outcome"
      ]
    }
  ]
}

MANDATORY RULES:
1. Identify ALL requirements first - be exhaustive
2. For EACH requirement, create EXACTLY 3 test cases
3. Formula: N requirements → N × 3 test cases
4. Each test case must have 3+ detailed steps"""

        user_prompt = f"""Analyze this test ticket and generate comprehensive test cases:

TICKET: {ticket.id}
SUMMARY: {ticket.summary}

DESCRIPTION:
{ticket.description[:2000]}

ACCEPTANCE CRITERIA:
"""
        for i, ac in enumerate(ticket.acceptance_criteria, 1):
            user_prompt += f"{i}. {ac}\n"

        # Add source ticket context if available
        if source_tickets_context:
            user_prompt += source_tickets_context

        # Add epic context if available (enhances context but not required for standalone mode)
        if epic_context_text:
            user_prompt += epic_context_text

        user_prompt += "\n\nGenerate test cases following the 3-per-requirement rule (Positive, Negative, Edge Case)."

        # Generate test cases with critic review
        result, critic_data = await asyncio.to_thread(
            generate_test_cases_with_retry,
            llm=llm_client,
            sys_prompt=sys_prompt,
            user_prompt=user_prompt,
            summary=ticket.summary,
            requirements_for_review=None,
            max_retries=2,
            progress_callback=send_progress_update
        )

        if not result:
            raise Exception("Failed to generate test cases after retries")

        requirements = result.get("requirements", [])
        test_cases_raw = result.get("test_cases", [])

        # Update ticket with test cases
        ticket.test_cases = test_cases_raw
        ticket.requirements = requirements
        ticket.analyzed = True

        await manager.send_progress({
            "type": "complete",
            "message": f"Generated {len(test_cases_raw)} test cases successfully"
        })

        return {
            "test_cases": test_cases_raw,
            "requirements": requirements,
            "ticket_info": {
                "id": ticket.id,
                "summary": ticket.summary,
                "requirements_count": len(requirements)
            },
            "generated_at": datetime.now().isoformat()
        }

    except Exception as e:
        await manager.send_progress({
            "type": "error",
            "message": str(e)
        })
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/test-cases/review-and-improve")
async def review_and_improve_test_cases(request: dict):
    """
    Review test cases with detailed analysis and provide improvement feedback
    """
    if not jira_client or not llm_client:
        raise HTTPException(status_code=401, detail="Not authenticated")

    try:
        test_cases = request.get('test_cases', [])
        ticket_context = request.get('ticket_context', {})
        requirements = request.get('requirements', [])

        if not test_cases:
            raise HTTPException(status_code=400, detail="No test cases provided")

        await manager.send_progress({
            "type": "progress",
            "step": "reviewing",
            "message": "Analyzing test cases for quality and completeness..."
        })

        # Build context for review
        summary = ticket_context.get('summary', 'Test Cases')
        description = ticket_context.get('description', '')

        # Step 1: Detailed Analysis
        analysis_prompt = """You are an expert QA test case reviewer. Analyze the provided test cases and provide detailed feedback.

Your analysis should include:
1. **Overall Quality Score** (0-100): Rate the overall quality of the test suite
2. **Coverage Analysis**: Identify which requirements/scenarios are well-covered and which are missing
3. **Specific Issues**: List specific problems found in individual test cases
4. **Missing Test Cases**: Suggest what types of test cases are missing
5. **Improvement Recommendations**: Provide actionable recommendations

Return your analysis in JSON format with these fields:
{
  "quality_score": <number 0-100>,
  "summary": "<brief overall assessment>",
  "strengths": ["<strength 1>", "<strength 2>", ...],
  "weaknesses": ["<weakness 1>", "<weakness 2>", ...],
  "missing_coverage": ["<missing area 1>", "<missing area 2>", ...],
  "recommendations": ["<recommendation 1>", "<recommendation 2>", ...]
}"""

        analysis_user_prompt = f"""Analyze these test cases:

TICKET: {summary}
DESCRIPTION: {description}

TEST CASES:
{json.dumps(test_cases, indent=2)}

REQUIREMENTS:
{json.dumps(requirements, indent=2) if requirements else 'None provided'}

Provide your detailed analysis."""

        analysis_result, error = llm_client.complete_json(analysis_prompt, analysis_user_prompt, max_tokens=2000)

        if error:
            raise Exception(f"Failed to analyze test cases: {error}")

        # Parse analysis
        analysis = json.loads(analysis_result) if isinstance(analysis_result, str) else analysis_result

        await manager.send_progress({
            "type": "complete",
            "message": f"Analysis complete! Quality score: {analysis.get('quality_score', 0)}/100"
        })

        # Return the review data in the format expected by frontend
        return {
            "success": True,
            "review": {
                "overall_score": analysis.get('quality_score', 0),
                "summary": analysis.get('summary', 'Analysis complete'),
                "strengths": analysis.get('strengths', []),
                "weaknesses": analysis.get('weaknesses', []),
                "missing_coverage": analysis.get('missing_coverage', []),
                "recommendations": analysis.get('recommendations', []),
                "test_case_count": len(test_cases)
            }
        }

    except Exception as e:
        await manager.send_progress({
            "type": "error",
            "message": str(e)
        })
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/test-cases/suggest-additional")
async def suggest_additional_test_cases(request: dict):
    """
    Generate improved test cases and/or additional test cases based on review feedback
    """
    if not jira_client or not llm_client:
        raise HTTPException(status_code=401, detail="Not authenticated")

    try:
        existing_test_cases = request.get('existing_test_cases', [])
        requirements = request.get('requirements', [])
        issues = request.get('issues', [])
        suggestions = request.get('suggestions', [])
        missing_scenarios = request.get('missingScenarios', [])

        if not existing_test_cases:
            raise HTTPException(status_code=400, detail="No existing test cases provided")

        await manager.send_progress({
            "type": "progress",
            "step": "improving",
            "message": "Generating improvements and additional test cases..."
        })

        # Build improvement prompt
        improvement_prompt = """You are an expert QA engineer. Based on the review feedback, you need to:
1. Improve existing test cases that have issues
2. Generate new test cases for missing scenarios

Return your response in JSON format with these fields:
{
  "improved_test_cases": [
    {
      "index": <original index of the test case being improved>,
      "id": "<test case ID>",
      "title": "<improved title>",
      "description": "<improved description>",
      "preconditions": "<improved preconditions>",
      "steps": ["Step 1: <action>", "Expected Result: <expected outcome>", "Step 2: <action>", "Expected Result: <expected outcome>", ...],
      "priority": "<priority>",
      "type": "<type>"
    }
  ],
  "new_test_cases": [
    {
      "id": "<new test case ID>",
      "title": "<title>",
      "description": "<description>",
      "preconditions": "<preconditions>",
      "steps": ["Step 1: <action>", "Expected Result: <expected outcome>", "Step 2: <action>", "Expected Result: <expected outcome>", ...],
      "priority": "<priority>",
      "type": "<type>"
    }
  ]
}

CRITICAL FORMATTING RULES:
- The "steps" array MUST alternate between "Step N:" and "Expected Result:" entries
- Each "Step N:" must be immediately followed by "Expected Result:" for that step
- Do NOT include a separate "expected_results" field
- Example: ["Step 1: Click login button", "Expected Result: Login form appears", "Step 2: Enter credentials", "Expected Result: Credentials are accepted"]
- Only include improved_test_cases if there are specific issues to fix in existing test cases
- Only include new_test_cases if there are missing scenarios to cover
- For improved test cases, include the original index so we know which test case to replace
- Generate realistic and actionable test cases"""

        user_prompt = f"""EXISTING TEST CASES:
{json.dumps(existing_test_cases, indent=2)}

REQUIREMENTS:
{json.dumps(requirements, indent=2) if requirements else 'None provided'}

ISSUES FOUND:
{json.dumps(issues, indent=2) if issues else 'None'}

IMPROVEMENT SUGGESTIONS:
{json.dumps(suggestions, indent=2) if suggestions else 'None'}

MISSING SCENARIOS:
{json.dumps(missing_scenarios, indent=2) if missing_scenarios else 'None'}

Based on this feedback, generate improved test cases and/or new test cases to address the issues and fill the gaps."""

        result, error = llm_client.complete_json(improvement_prompt, user_prompt, max_tokens=4000)

        if error:
            raise Exception(f"Failed to generate improvements: {error}")

        # Parse result
        improvements = json.loads(result) if isinstance(result, str) else result

        improved_cases = improvements.get('improved_test_cases', [])
        new_cases = improvements.get('new_test_cases', [])

        await manager.send_progress({
            "type": "complete",
            "message": f"Generated {len(improved_cases)} improvements and {len(new_cases)} new test cases"
        })

        return {
            "success": True,
            "improved_test_cases": improved_cases,
            "new_test_cases": new_cases,
            "total_improvements": len(improved_cases),
            "total_new": len(new_cases)
        }

    except Exception as e:
        await manager.send_progress({
            "type": "error",
            "message": str(e)
        })
        raise HTTPException(status_code=500, detail=str(e))


# ============================================================================
# WebSocket for Real-time Progress
# ============================================================================

@app.websocket("/ws/progress")
async def websocket_endpoint(websocket: WebSocket):
    """WebSocket endpoint for real-time progress updates."""
    await manager.connect(websocket)
    try:
        while True:
            # Keep connection alive
            data = await websocket.receive_text()
            # Echo back for heartbeat
            await websocket.send_json({"type": "heartbeat"})
    except WebSocketDisconnect:
        manager.disconnect(websocket)
    except Exception as e:
        print(f"WebSocket error: {e}")
        manager.disconnect(websocket)


# ============================================================================
# Settings Management
# ============================================================================

@app.get("/api/settings")
async def get_settings():
    """Get user settings"""
    return user_settings


@app.post("/api/settings")
async def save_settings(settings: dict):
    """Save user settings"""
    global user_settings
    user_settings.update(settings)
    return {"success": True, "settings": user_settings}


# ============================================================================
# Cache Management
# ============================================================================

@app.get("/api/cache/stats")
async def get_cache_stats():
    """
    Get cache statistics including hit rate, memory usage, and performance metrics.

    Returns detailed statistics about the LLM response cache.
    """
    if not llm_client:
        raise HTTPException(status_code=400, detail="LLM client not initialized")

    try:
        stats = llm_client.cache_client.get_stats()
        return {
            "success": True,
            "stats": stats
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to get cache stats: {str(e)}")


@app.post("/api/cache/clear")
async def clear_cache(pattern: Optional[str] = None):
    """
    Clear cached LLM responses.

    Args:
        pattern: Optional pattern to match cache keys (Redis only)

    Returns:
        Number of cache entries cleared
    """
    if not llm_client:
        raise HTTPException(status_code=400, detail="LLM client not initialized")

    try:
        cleared_count = llm_client.cache_client.clear(pattern=pattern)
        return {
            "success": True,
            "cleared_count": cleared_count,
            "message": f"Cleared {cleared_count} cache entries"
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to clear cache: {str(e)}")


@app.delete("/api/cache/clear/{pattern}")
async def clear_cache_by_pattern(pattern: str):
    """
    Clear cached LLM responses matching a specific pattern.

    Args:
        pattern: Pattern to match cache keys (e.g., "llm_cache:v1:*")

    Returns:
        Number of cache entries cleared
    """
    if not llm_client:
        raise HTTPException(status_code=400, detail="LLM client not initialized")

    try:
        cleared_count = llm_client.cache_client.clear(pattern=pattern)
        return {
            "success": True,
            "cleared_count": cleared_count,
            "pattern": pattern,
            "message": f"Cleared {cleared_count} cache entries matching pattern '{pattern}'"
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to clear cache: {str(e)}")


# ============================================================================
# Health Check
# ============================================================================

@app.get("/")
async def root():
    """Health check endpoint."""
    return {
        "name": "AI Tester Framework API",
        "version": "3.0.0",
        "status": "running"
    }


@app.get("/api/health")
async def health_check():
    """Detailed health check."""
    return {
        "status": "healthy",
        "authenticated": jira_client is not None and llm_client is not None,
        "timestamp": datetime.now().isoformat()
    }


if __name__ == "__main__":
    import uvicorn
    # Security: Use environment variable for host binding, default to localhost
    # Set API_HOST=0.0.0.0 in production environment if needed
    host = os.getenv("API_HOST", "127.0.0.1")
    uvicorn.run(app, host=host, port=8000)
# Trigger reload
